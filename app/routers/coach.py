from typing import Annotated, Optional
from datetime import date
from io import BytesIO
import re
from urllib.parse import quote
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func, delete
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app.models.enums import AttendanceStatus, StudentStatus
from app.core.db import get_db
from app.core.permissions import PERM_ATTENDANCE_COACH_MARK
from app.models.domain import (
    Group,
    Student,
    GroupStatus,
    GroupPerformanceTable,
    GroupPerformanceMatch,
    GroupPerformanceCell,
)
from app.models.attendance import Session, Attendance
from app.schemas.group import GroupRead
from app.schemas.attendance import (
    SessionRead, SessionCreate, AttendanceCreate, SessionWithAttendances,
    BulkAttendanceCreate, AttendanceStats, AttendanceRead
)
from app.schemas.performance_table import (
    PerformanceTableRead,
    PerformanceTableUpsert,
    PerformanceMatchRead,
    PerformanceRowRead,
    PerformanceColumnCreate,
    PerformanceColumnUpdate,
    PerformanceColumnReorder,
)
from app.schemas.common import DataResponse
from app.deps import require_permission, CurrentUser
from app.services.file_upload import file_upload_service

router = APIRouter(prefix="/coach", tags=["Coach"])


async def _get_coach_group_or_404(db: AsyncSession, user_id: int, group_id: int) -> Group:
    group_result = await db.execute(
        select(Group)
        .options(selectinload(Group.coach))
        .where(
            Group.id == group_id,
            Group.coach_id == user_id,
            Group.status != GroupStatus.DELETED,
        )
    )
    group = group_result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found or not assigned to you")
    return group


async def _build_performance_table_response(
    db: AsyncSession,
    group: Group,
    season_year: int,
    table: GroupPerformanceTable | None,
) -> PerformanceTableRead:
    students_result = await db.execute(
        select(Student)
        .where(
            Student.group_id == group.id,
            Student.status != StudentStatus.DELETED,
        )
        .order_by(Student.last_name.asc(), Student.first_name.asc())
    )
    students = students_result.scalars().all()

    matches = table.matches if table else []
    match_reads = [
        PerformanceMatchRead(
            id=match.id,
            position=match.position,
            tour_label=match.tour_label,
            match_date=match.match_date,
            opponent=match.opponent,
        )
        for match in matches
    ]

    cell_map = {}
    if table:
        cell_map = {(cell.student_id, cell.match_id): cell.value for cell in table.cells}

    rows = []
    for student in students:
        rows.append(
            PerformanceRowRead(
                student_id=student.id,
                student_name=f"{student.last_name} {student.first_name}",
                millati=student.millati,
                cells=[cell_map.get((student.id, match.id)) for match in matches],
            )
        )

    coach_name = None
    if group.coach:
        coach_name = group.coach.full_name or group.coach.phone or str(group.coach.id)

    return PerformanceTableRead(
        table_id=table.id if table else None,
        group_id=group.id,
        group_name=group.name,
        coach_id=group.coach_id,
        coach_name=coach_name,
        title=table.title if table else f"{group.name} - {season_year}",
        season_year=season_year,
        matches=match_reads,
        rows=rows,
        updated_at=table.updated_at if table else None,
    )


async def _get_performance_table_for_group(
    db: AsyncSession,
    group_id: int,
    season_year: int,
) -> GroupPerformanceTable | None:
    table_result = await db.execute(
        select(GroupPerformanceTable)
        .execution_options(populate_existing=True)
        .options(
            selectinload(GroupPerformanceTable.matches),
            selectinload(GroupPerformanceTable.cells),
            selectinload(GroupPerformanceTable.group).selectinload(Group.coach),
        )
        .where(
            GroupPerformanceTable.group_id == group_id,
            GroupPerformanceTable.season_year == season_year,
        )
    )
    return table_result.scalar_one_or_none()


async def _validate_performance_values(
    db: AsyncSession,
    group_id: int,
    values: list,
) -> None:
    if not values:
        return

    student_ids = [item.student_id for item in values]
    if len(set(student_ids)) != len(student_ids):
        raise HTTPException(status_code=400, detail="Duplicate student_id in column values")

    students_result = await db.execute(
        select(Student.id).where(
            Student.group_id == group_id,
            Student.status != StudentStatus.DELETED,
            Student.id.in_(student_ids),
        )
    )
    valid_ids = set(students_result.scalars().all())
    invalid_ids = sorted(set(student_ids) - valid_ids)
    if invalid_ids:
        raise HTTPException(status_code=400, detail=f"Invalid student_ids for this group: {invalid_ids}")


@router.get("/groups", response_model=DataResponse[list[GroupRead]], dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))])
async def get_coach_groups(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Get coach's assigned groups (excluding deleted groups)"""
    result = await db.execute(
        select(Group).where(
            Group.coach_id == user.id,
            Group.status != GroupStatus.DELETED
        )
    )
    groups = result.scalars().all()
    return DataResponse(data=[GroupRead.model_validate(g) for g in groups])


@router.get(
    "/groups/{group_id}/performance-table",
    response_model=DataResponse[PerformanceTableRead],
    dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))],
)
async def get_group_performance_table(
    group_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    season_year: Optional[int] = Query(None, description="Season year, defaults to current year"),
):
    if season_year is None:
        season_year = date.today().year

    group = await _get_coach_group_or_404(db, user.id, group_id)
    table_result = await db.execute(
        select(GroupPerformanceTable)
        .options(
            selectinload(GroupPerformanceTable.matches),
            selectinload(GroupPerformanceTable.cells),
            selectinload(GroupPerformanceTable.group).selectinload(Group.coach),
        )
        .where(
            GroupPerformanceTable.group_id == group_id,
            GroupPerformanceTable.season_year == season_year,
        )
    )
    table = table_result.scalar_one_or_none()
    return DataResponse(data=await _build_performance_table_response(db, group, season_year, table))


@router.put(
    "/groups/{group_id}/performance-table",
    response_model=DataResponse[PerformanceTableRead],
    dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))],
)
async def save_group_performance_table(
    group_id: int,
    data: PerformanceTableUpsert,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    group = await _get_coach_group_or_404(db, user.id, group_id)

    students_result = await db.execute(
        select(Student.id)
        .where(
            Student.group_id == group_id,
            Student.status != StudentStatus.DELETED,
        )
    )
    valid_student_ids = set(students_result.scalars().all())

    if len({row.student_id for row in data.rows}) != len(data.rows):
        raise HTTPException(status_code=400, detail="Duplicate student_id in rows payload")

    invalid_student_ids = sorted({row.student_id for row in data.rows} - valid_student_ids)
    if invalid_student_ids:
        raise HTTPException(status_code=400, detail=f"Invalid student_ids for this group: {invalid_student_ids}")

    match_count = len(data.matches)
    for row in data.rows:
        if len(row.cells) != match_count:
            raise HTTPException(
                status_code=400,
                detail=f"Row for student_id {row.student_id} must have exactly {match_count} cells",
            )

    table_result = await db.execute(
        select(GroupPerformanceTable)
        .options(
            selectinload(GroupPerformanceTable.matches),
            selectinload(GroupPerformanceTable.cells),
            selectinload(GroupPerformanceTable.group).selectinload(Group.coach),
        )
        .where(
            GroupPerformanceTable.group_id == group_id,
            GroupPerformanceTable.season_year == data.season_year,
        )
    )
    table = table_result.scalar_one_or_none()

    if not table:
        table = GroupPerformanceTable(
            group_id=group_id,
            season_year=data.season_year,
            title=data.title or f"{group.name} - {data.season_year}",
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
        db.add(table)
        await db.flush()
    else:
        table.title = data.title or f"{group.name} - {data.season_year}"
        table.updated_by_user_id = user.id
        await db.execute(delete(GroupPerformanceCell).where(GroupPerformanceCell.table_id == table.id))
        await db.execute(delete(GroupPerformanceMatch).where(GroupPerformanceMatch.table_id == table.id))
        await db.flush()

    created_matches: list[GroupPerformanceMatch] = []
    for index, match_data in enumerate(data.matches, start=1):
        match = GroupPerformanceMatch(
            table_id=table.id,
            position=index,
            tour_label=match_data.tour_label,
            match_date=match_data.match_date,
            opponent=match_data.opponent,
        )
        db.add(match)
        created_matches.append(match)

    await db.flush()

    for row in data.rows:
        for index, raw_value in enumerate(row.cells):
            value = raw_value if raw_value not in ("", None) else None
            if value is None:
                continue
            db.add(
                GroupPerformanceCell(
                    table_id=table.id,
                    match_id=created_matches[index].id,
                    student_id=row.student_id,
                    value=value,
                )
            )

    await db.commit()
    db.expire_all()
    refreshed_table = await _get_performance_table_for_group(db, group_id, data.season_year)
    return DataResponse(data=await _build_performance_table_response(db, group, data.season_year, refreshed_table))


@router.post(
    "/groups/{group_id}/performance-table/columns",
    response_model=DataResponse[PerformanceTableRead],
    dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))],
)
async def add_performance_table_column(
    group_id: int,
    data: PerformanceColumnCreate,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    group = await _get_coach_group_or_404(db, user.id, group_id)
    await _validate_performance_values(db, group_id, data.values)

    table = await _get_performance_table_for_group(db, group_id, data.season_year)
    if not table:
        table = GroupPerformanceTable(
            group_id=group_id,
            season_year=data.season_year,
            title=f"{group.name} - {data.season_year}",
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
        db.add(table)
        await db.flush()

    next_position = (max((match.position for match in table.matches), default=0) + 1) if table.matches else 1
    match = GroupPerformanceMatch(
        table_id=table.id,
        position=next_position,
        tour_label=data.tour_label,
        match_date=data.match_date,
        opponent=data.opponent,
    )
    db.add(match)
    await db.flush()

    for item in data.values:
        if item.value in ("", None):
            continue
        db.add(
            GroupPerformanceCell(
                table_id=table.id,
                match_id=match.id,
                student_id=item.student_id,
                value=item.value,
            )
        )

    table.updated_by_user_id = user.id
    await db.commit()
    db.expire_all()
    refreshed_table = await _get_performance_table_for_group(db, group_id, data.season_year)
    return DataResponse(data=await _build_performance_table_response(db, group, data.season_year, refreshed_table))


@router.patch(
    "/groups/{group_id}/performance-table/columns/{match_id}",
    response_model=DataResponse[PerformanceTableRead],
    dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))],
)
async def update_performance_table_column(
    group_id: int,
    match_id: int,
    data: PerformanceColumnUpdate,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    group = await _get_coach_group_or_404(db, user.id, group_id)
    if data.values is not None:
        await _validate_performance_values(db, group_id, data.values)

    match_result = await db.execute(
        select(GroupPerformanceMatch)
        .options(selectinload(GroupPerformanceMatch.table))
        .join(GroupPerformanceTable)
        .where(
            GroupPerformanceMatch.id == match_id,
            GroupPerformanceTable.group_id == group_id,
        )
    )
    match = match_result.scalar_one_or_none()
    if not match:
        raise HTTPException(status_code=404, detail="Column not found for this group")

    match.tour_label = data.tour_label
    match.match_date = data.match_date
    match.opponent = data.opponent
    match.table.updated_by_user_id = user.id

    if data.values is not None:
        await db.execute(delete(GroupPerformanceCell).where(GroupPerformanceCell.match_id == match_id))
        for item in data.values:
            if item.value in ("", None):
                continue
            db.add(
                GroupPerformanceCell(
                    table_id=match.table_id,
                    match_id=match.id,
                    student_id=item.student_id,
                    value=item.value,
                )
            )

    await db.commit()
    db.expire_all()
    refreshed_table = await _get_performance_table_for_group(db, group_id, match.table.season_year)
    return DataResponse(data=await _build_performance_table_response(db, group, match.table.season_year, refreshed_table))


@router.delete(
    "/groups/{group_id}/performance-table/columns/{match_id}",
    response_model=DataResponse[PerformanceTableRead],
    dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))],
)
async def delete_performance_table_column(
    group_id: int,
    match_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    group = await _get_coach_group_or_404(db, user.id, group_id)

    match_result = await db.execute(
        select(GroupPerformanceMatch)
        .options(selectinload(GroupPerformanceMatch.table).selectinload(GroupPerformanceTable.matches))
        .join(GroupPerformanceTable)
        .where(
            GroupPerformanceMatch.id == match_id,
            GroupPerformanceTable.group_id == group_id,
        )
    )
    match = match_result.scalar_one_or_none()
    if not match:
        raise HTTPException(status_code=404, detail="Column not found for this group")

    season_year = match.table.season_year
    table_id = match.table_id

    await db.delete(match)
    await db.flush()

    remaining_matches_result = await db.execute(
        select(GroupPerformanceMatch)
        .where(GroupPerformanceMatch.table_id == table_id)
        .order_by(GroupPerformanceMatch.position.asc(), GroupPerformanceMatch.id.asc())
    )
    for index, remaining_match in enumerate(remaining_matches_result.scalars().all(), start=1):
        remaining_match.position = index

    match.table.updated_by_user_id = user.id
    await db.commit()
    db.expire_all()
    refreshed_table = await _get_performance_table_for_group(db, group_id, season_year)
    return DataResponse(data=await _build_performance_table_response(db, group, season_year, refreshed_table))


@router.patch(
    "/groups/{group_id}/performance-table/columns-reorder",
    response_model=DataResponse[PerformanceTableRead],
    dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))],
)
async def reorder_performance_table_columns(
    group_id: int,
    data: PerformanceColumnReorder,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    group = await _get_coach_group_or_404(db, user.id, group_id)
    table = await _get_performance_table_for_group(db, group_id, data.season_year)
    if not table:
        raise HTTPException(status_code=404, detail="Performance table not found for this season")

    current_match_ids = [match.id for match in table.matches]
    if set(current_match_ids) != set(data.match_ids) or len(current_match_ids) != len(data.match_ids):
        raise HTTPException(status_code=400, detail="match_ids must exactly match the table's current columns")

    match_map = {match.id: match for match in table.matches}
    offset = len(data.match_ids) + 1000
    for match in table.matches:
        match.position = match.position + offset
    await db.flush()

    for index, match_id in enumerate(data.match_ids, start=1):
        match_map[match_id].position = index

    table.updated_by_user_id = user.id
    await db.commit()
    db.expire_all()
    refreshed_table = await _get_performance_table_for_group(db, group_id, data.season_year)
    return DataResponse(data=await _build_performance_table_response(db, group, data.season_year, refreshed_table))


@router.get(
    "/groups/{group_id}/performance-table/export",
    dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))],
)
async def export_group_performance_table(
    group_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    season_year: Optional[int] = Query(None, description="Season year, defaults to current year"),
):
    if season_year is None:
        season_year = date.today().year

    group = await _get_coach_group_or_404(db, user.id, group_id)
    table_result = await db.execute(
        select(GroupPerformanceTable)
        .options(
            selectinload(GroupPerformanceTable.matches),
            selectinload(GroupPerformanceTable.cells),
            selectinload(GroupPerformanceTable.group).selectinload(Group.coach),
        )
        .where(
            GroupPerformanceTable.group_id == group_id,
            GroupPerformanceTable.season_year == season_year,
        )
    )
    table = table_result.scalar_one_or_none()
    table_data = await _build_performance_table_response(db, group, season_year, table)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Performance Table"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    title_font = Font(bold=True, size=13)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    total_columns = 3 + len(table_data.matches)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(total_columns, 3))
    title_cell = sheet.cell(row=1, column=1, value=table_data.title)
    title_cell.font = title_font
    title_cell.alignment = center

    sheet.cell(row=2, column=1, value="No")
    sheet.cell(row=2, column=2, value="F.I.SH")
    sheet.cell(row=2, column=3, value="Millati")

    for col in range(1, 4):
        sheet.cell(row=2, column=col).font = header_font
        sheet.cell(row=2, column=col).fill = header_fill
        sheet.cell(row=2, column=col).alignment = center
        sheet.cell(row=2, column=col).border = border

    for index, match in enumerate(table_data.matches, start=4):
        header_parts = []
        if match.tour_label:
            header_parts.append(match.tour_label)
        if match.match_date:
            header_parts.append(match.match_date.strftime("%d.%m"))
        if match.opponent:
            header_parts.append(match.opponent)
        header_value = "\n".join(header_parts) if header_parts else f"Match {index - 3}"
        cell = sheet.cell(row=2, column=index, value=header_value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    sheet.row_dimensions[2].height = 60

    for row_index, row in enumerate(table_data.rows, start=3):
        sheet.cell(row=row_index, column=1, value=row_index - 2)
        sheet.cell(row=row_index, column=2, value=row.student_name)
        sheet.cell(row=row_index, column=3, value=row.millati or "")
        for col in range(1, 4):
            sheet.cell(row=row_index, column=col).border = border
            sheet.cell(row=row_index, column=col).alignment = Alignment(vertical="center")
        for cell_index, value in enumerate(row.cells, start=4):
            data_cell = sheet.cell(row=row_index, column=cell_index, value=value or "")
            data_cell.alignment = center
            data_cell.border = border

    sheet.column_dimensions["A"].width = 6
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 16
    for index in range(4, total_columns + 1):
        column_letter = sheet.cell(row=2, column=index).column_letter
        sheet.column_dimensions[column_letter].width = 14

    file_buffer = BytesIO()
    workbook.save(file_buffer)
    file_buffer.seek(0)

    safe_group = re.sub(r"[^A-Za-z0-9._-]", "_", table_data.group_name)
    filename = f"performance_table_{safe_group}_{season_year}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        file_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"{filename}\"; "
                f"filename*=UTF-8''{encoded_filename}"
            )
        },
    )


@router.get("/sessions", response_model=DataResponse[list[SessionRead]], dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))])
async def get_coach_sessions(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_filter: Optional[date] = Query(None, alias="date"),
    group_id: Optional[int] = Query(None),
):
    """
    Get sessions for coach's groups (excluding deleted groups).
    Coach can view sessions by date and group, and mark attendance.
    """
    session_date = date_filter or date.today()

    query = select(Session).join(Group).where(
        and_(
            Group.coach_id == user.id,
            Session.session_date == session_date,
            Group.status != GroupStatus.DELETED
        )
    )

    if group_id:
        query = query.where(Session.group_id == group_id)

    result = await db.execute(query)
    sessions = result.scalars().all()
    return DataResponse(data=[SessionRead.model_validate(s) for s in sessions])


@router.post("/sessions/{session_id}/attendance", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))])
async def mark_attendance(
    session_id: int,
    data: AttendanceCreate,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    session_result = await db.execute(select(Session).where(Session.id == session_id))
    session = session_result.scalar_one_or_none()

    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    # Check if attendance has already been marked for this session
    existing_attendance = await db.execute(
        select(Attendance).where(Attendance.session_id == session_id).limit(1)
    )
    if existing_attendance.scalar_one_or_none():
        raise HTTPException(
            status_code=400,
            detail="Bu dars uchun davomat allaqachon qilib bo'lingan. Bir darsga faqat bir marta davomat qilish mumkin."
        )

    student_result = await db.execute(select(Student).where(Student.id == data.student_id))
    if not student_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail=f"Student with ID {data.student_id} not found")

    attendance = Attendance(
        session_id=session_id,
        student_id=data.student_id,
        status=data.status,
        comment=data.comment,
        marked_by_user_id=user.id,
    )
    db.add(attendance)
    await db.commit()
    await db.refresh(attendance)

    return DataResponse(data={"message": "Attendance marked successfully", "attendance_id": attendance.id})




@router.get("/sessions/{session_id}", response_model=DataResponse[SessionWithAttendances], dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))])
async def get_session_details(
    session_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Get session details with all attendance records"""
    session_result = await db.execute(
        select(Session)
        .options(selectinload(Session.attendances))
        .join(Group)
        .where(Session.id == session_id, Group.coach_id == user.id)
    )
    session = session_result.scalar_one_or_none()

    if not session:
        raise HTTPException(
            status_code=404,
            detail="Session not found or you do not have permission to view it"
        )

    return DataResponse(data=SessionWithAttendances.model_validate(session))


@router.post("/sessions/{session_id}/bulk-attendance", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))])
async def mark_bulk_attendance(
    session_id: int,
    data: BulkAttendanceCreate,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Mark attendance for multiple students at once"""
    # Verify session exists and coach has access
    session_result = await db.execute(
        select(Session)
        .join(Group)
        .where(Session.id == session_id, Group.coach_id == user.id)
    )
    session = session_result.scalar_one_or_none()

    if not session:
        raise HTTPException(
            status_code=404,
            detail="Session not found or you do not have permission to mark attendance"
        )

    # Check if attendance has already been marked for this session
    existing_attendance = await db.execute(
        select(Attendance).where(Attendance.session_id == session_id).limit(1)
    )
    if existing_attendance.scalar_one_or_none():
        raise HTTPException(
            status_code=400,
            detail="Bu dars uchun davomat allaqachon qilib bo'lingan. Bir darsga faqat bir marta davomat qilish mumkin."
        )

    if data.session_id != session_id:
        raise HTTPException(status_code=400, detail="Session ID mismatch")

    # Create attendance records
    attendance_records = []
    for att_data in data.attendances:
        # Verify student exists
        student_result = await db.execute(select(Student).where(Student.id == att_data.student_id))
        if not student_result.scalar_one_or_none():
            raise HTTPException(status_code=404, detail=f"Student with ID {att_data.student_id} not found")

        attendance = Attendance(
            session_id=session_id,
            student_id=att_data.student_id,
            status=att_data.status,
            comment=att_data.comment,
            marked_by_user_id=user.id,
        )
        db.add(attendance)
        attendance_records.append(attendance)

    await db.commit()

    return DataResponse(data={
        "message": f"Successfully marked attendance for {len(attendance_records)} students",
        "count": len(attendance_records)
    })


@router.get("/groups/{group_id}/attendance-stats", response_model=DataResponse[AttendanceStats], dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))])
async def get_group_attendance_stats(
    group_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
):
    """Get attendance statistics for a group"""
    # Verify group exists and coach has access
    group_result = await db.execute(
        select(Group).where(Group.id == group_id, Group.coach_id == user.id)
    )
    group = group_result.scalar_one_or_none()

    if not group:
        raise HTTPException(
            status_code=404,
            detail="Group not found or you do not have permission to view statistics"
        )

    # Get all sessions for the group
    sessions_query = select(Session).where(Session.group_id == group_id)
    if from_date:
        sessions_query = sessions_query.where(Session.session_date >= from_date)
    if to_date:
        sessions_query = sessions_query.where(Session.session_date <= to_date)

    sessions_result = await db.execute(sessions_query)
    sessions = sessions_result.scalars().all()
    session_ids = [s.id for s in sessions]

    if not session_ids:
        return DataResponse(data=AttendanceStats(
            total_sessions=0,
            present_count=0,
            absent_count=0,
            late_count=0,
            attendance_rate=0.0
        ))

    # Count attendance by status
    present_count = await db.execute(
        select(func.count(Attendance.id)).where(
            Attendance.session_id.in_(session_ids),
            Attendance.status == AttendanceStatus.PRESENT
        )
    )
    absent_count = await db.execute(
        select(func.count(Attendance.id)).where(
            Attendance.session_id.in_(session_ids),
            Attendance.status == AttendanceStatus.ABSENT
        )
    )
    late_count = await db.execute(
        select(func.count(Attendance.id)).where(
            Attendance.session_id.in_(session_ids),
            Attendance.status == AttendanceStatus.LATE
        )
    )

    present = present_count.scalar() or 0
    absent = absent_count.scalar() or 0
    late = late_count.scalar() or 0
    total = present + absent + late

    attendance_rate = (present / total * 100) if total > 0 else 0.0

    return DataResponse(data=AttendanceStats(
        total_sessions=len(sessions),
        present_count=present,
        absent_count=absent,
        late_count=late,
        attendance_rate=round(attendance_rate, 2)
    ))


@router.get("/students/{student_id}/attendance-stats", response_model=DataResponse[AttendanceStats], dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))])
async def get_student_attendance_stats(
    student_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
):
    """Get attendance statistics for a specific student"""
    # Verify student exists and is in one of coach's groups
    student_result = await db.execute(
        select(Student)
        .join(Group)
        .where(Student.id == student_id, Group.coach_id == user.id)
    )
    student = student_result.scalar_one_or_none()

    if not student:
        raise HTTPException(
            status_code=404,
            detail="Student not found or not in your groups"
        )

    # Get attendance records for this student
    attendance_query = select(Attendance).where(Attendance.student_id == student_id)

    if from_date or to_date:
        attendance_query = attendance_query.join(Session)
        if from_date:
            attendance_query = attendance_query.where(Session.session_date >= from_date)
        if to_date:
            attendance_query = attendance_query.where(Session.session_date <= to_date)

    attendances_result = await db.execute(attendance_query)
    attendances = attendances_result.scalars().all()

    present = sum(1 for a in attendances if a.status == AttendanceStatus.PRESENT)
    absent = sum(1 for a in attendances if a.status == AttendanceStatus.ABSENT)
    late = sum(1 for a in attendances if a.status == AttendanceStatus.LATE)
    total = len(attendances)

    attendance_rate = (present / total * 100) if total > 0 else 0.0

    # Get total sessions count
    sessions_query = select(func.count(Session.id)).join(Group).where(
        Group.id == student.group_id
    )
    if from_date:
        sessions_query = sessions_query.where(Session.session_date >= from_date)
    if to_date:
        sessions_query = sessions_query.where(Session.session_date <= to_date)

    total_sessions_result = await db.execute(sessions_query)
    total_sessions = total_sessions_result.scalar() or 0

    return DataResponse(data=AttendanceStats(
        total_sessions=total_sessions,
        present_count=present,
        absent_count=absent,
        late_count=late,
        attendance_rate=round(attendance_rate, 2)
    ))


@router.get("/my-attendances", response_model=DataResponse[list[AttendanceRead]], dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))])
async def get_coach_created_attendances(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    group_id: Optional[int] = None,
    student_id: Optional[int] = None,
):
    """
    Get all attendance records created by the coach (read-only).

    This endpoint allows coaches to view all attendances they have marked,
    including for all groups and students. Coaches cannot edit these records
    through this endpoint - it's read-only.

    Filters:
    - from_date: Filter by session date (start)
    - to_date: Filter by session date (end)
    - group_id: Filter by specific group
    - student_id: Filter by specific student
    """
    # Build query for attendances created by this coach
    attendance_query = select(Attendance).options(
        selectinload(Attendance.student),
        selectinload(Attendance.session).selectinload(Session.group),
        selectinload(Attendance.marked_by)
    ).where(Attendance.marked_by_user_id == user.id)

    # Apply date filters via session
    if from_date or to_date:
        attendance_query = attendance_query.join(Session)
        if from_date:
            attendance_query = attendance_query.where(Session.session_date >= from_date)
        if to_date:
            attendance_query = attendance_query.where(Session.session_date <= to_date)

    # Apply group filter via session
    if group_id:
        if not (from_date or to_date):  # Only join if not already joined
            attendance_query = attendance_query.join(Session)
        attendance_query = attendance_query.where(Session.group_id == group_id)

    # Apply student filter
    if student_id:
        attendance_query = attendance_query.where(Attendance.student_id == student_id)

    # Order by most recent first
    attendance_query = attendance_query.order_by(Attendance.created_at.desc())

    attendances_result = await db.execute(attendance_query)
    attendances = attendances_result.scalars().all()

    return DataResponse(data=[AttendanceRead.model_validate(a) for a in attendances])


@router.post("/sessions/{session_id}/upload-konspekt", response_model=DataResponse[SessionRead], dependencies=[Depends(require_permission(PERM_ATTENDANCE_COACH_MARK))])
async def upload_konspekt(
    session_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    file: UploadFile = File(...),
    description: Optional[str] = Form(None),
):
    """
    Upload konspekt document to S3 and update session.
    Accepts: .pdf, .docx, .doc, images (.jpg, .jpeg, .png)

    Process:
    1. Coach marks attendance
    2. Coach uploads konspekt file with optional description (this endpoint)
    3. File is uploaded to S3 and session is updated with URL and description

    Form Data:
    - file: konspekt file (required)
    - description: optional description about the session
    """
    if not file_upload_service:
        raise HTTPException(
            status_code=500,
            detail="S3 service not configured. Please add AWS credentials to .env file: AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY, AWS_BUCKET_NAME"
        )

    # Verify session exists and coach has access to it
    session_result = await db.execute(
        select(Session)
        .join(Group)
        .where(Session.id == session_id, Group.coach_id == user.id, Group.status != GroupStatus.DELETED)
    )
    session = session_result.scalar_one_or_none()

    if not session:
        raise HTTPException(
            status_code=404,
            detail="Session not found or you do not have permission to upload konspekt"
        )

    # Validate file type
    allowed_extensions = ['.pdf', '.docx', '.doc', '.jpg', '.jpeg', '.png']
    file_ext = '.' + file.filename.split('.')[-1].lower() if '.' in file.filename else ''

    if file_ext not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type. Allowed: {', '.join(allowed_extensions)}"
        )

    try:
        # Upload to S3 in konspekts folder
        # Note: Bucket policy must allow public read for konspekts/*
        s3_url = await file_upload_service.upload_file(
            file=file,
            prefix="konspekts",
            make_public=True
        )

        # Update session with konspekt URL and description
        session.konspekt_url = s3_url
        if description is not None:
            session.description = description

        await db.commit()
        await db.refresh(session)

        return DataResponse(data=SessionRead.model_validate(session))
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to upload file: {str(e)}"
        )
