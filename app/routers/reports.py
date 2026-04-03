from typing import Annotated, Optional
from datetime import date, datetime, timedelta
from decimal import Decimal
import json
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import Date, select, func, and_, cast, or_
from sqlalchemy.dialects.postgresql import JSONB
from dateutil.relativedelta import relativedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from app.core.db import get_db
from app.core.permissions import PERM_REPORTS_DASHBOARD_VIEW, PERM_REPORTS_FINANCE_VIEW, PERM_REPORTS_ATTENDANCE_VIEW
from app.models.domain import Student, Group
from app.models.finance import Transaction
from app.models.attendance import Session, Attendance
from app.models.enums import StudentStatus, PaymentStatus, AttendanceStatus, ContractStatus, PaymentSettlementType, PaymentSource
from app.schemas.report import (
    DashboardPeriodSummary,
    DashboardSourceBreakdown,
    DashboardSummary,
    DashboardTrendPoint,
    FinanceReport,
    FinanceReportItem,
    GroupAttendanceReport,
    StudentAttendanceReport,
    DebtorMonthItem,
    DebtorItem,
    PayerItem,
)
from app.schemas.common import DataResponse, PaginationMeta
from app.deps import require_permission
from app.services.transaction_reporting import build_month_range, allocate_amount_for_period

router = APIRouter(prefix="/reports", tags=["Reports"])


def _months_elapsed(start_date: date, end_date: date) -> int:
    months = 0
    current_date = start_date
    while current_date <= end_date:
        months += 1
        current_date = current_date + relativedelta(months=1)
    return months


def _build_excel_file(sheet_name: str, headers: list[str], rows: list[list[object]]) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)

    for row in rows:
        worksheet.append(row)

    for header_cell in worksheet[1]:
        header_cell.font = Font(bold=True)

    for column_index, header in enumerate(headers, 1):
        column_letter = get_column_letter(column_index)
        max_length = len(str(header))
        for cell in worksheet[column_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _to_decimal(value: object) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if value is None:
        return Decimal("0")
    return Decimal(str(value))


def _to_money(amount: Decimal) -> float:
    return float(amount.quantize(Decimal("0.01")))


def _format_month_label(year: int, month: int) -> str:
    return f"{year}-{month:02d}"


def _resolve_source_bucket(source: object) -> str:
    if source == PaymentSource.PAYME or str(source).lower() == PaymentSource.PAYME.value:
        return "payme"
    if source == PaymentSource.CLICK or str(source).lower() == PaymentSource.CLICK.value:
        return "click"
    return "other"


def _build_dashboard_period_summary(
    period_from: date,
    period_to: date,
    rows: list,
) -> DashboardPeriodSummary:
    day_count = (period_to - period_from).days + 1
    days = [period_from + timedelta(days=offset) for offset in range(max(day_count, 0))]

    daily_totals: dict[date, dict[str, Decimal]] = {
        day: {
            "inflow": Decimal("0"),
            "outflow": Decimal("0"),
            "net_amount": Decimal("0"),
            "payme": Decimal("0"),
            "click": Decimal("0"),
            "other": Decimal("0"),
        }
        for day in days
    }

    total_inflow = Decimal("0")
    total_outflow = Decimal("0")
    net_amount = Decimal("0")
    successful_transactions = 0
    source_totals: dict[str, dict[str, Decimal | int]] = {
        "payme": {"amount": Decimal("0"), "transaction_count": 0},
        "click": {"amount": Decimal("0"), "transaction_count": 0},
        "other": {"amount": Decimal("0"), "transaction_count": 0},
    }

    for row in rows:
        paid_date = row.paid_date
        if paid_date is None or paid_date not in daily_totals:
            continue

        amount = _to_decimal(row.amount)
        source_bucket = _resolve_source_bucket(row.source)
        day_entry = daily_totals[paid_date]
        source_entry = source_totals[source_bucket]

        successful_transactions += 1
        source_entry["transaction_count"] = int(source_entry["transaction_count"]) + 1
        day_entry["net_amount"] += amount
        net_amount += amount

        if amount >= 0:
            day_entry["inflow"] += amount
            day_entry[source_bucket] += amount
            total_inflow += amount
            source_entry["amount"] = _to_decimal(source_entry["amount"]) + amount
        else:
            outflow = -amount
            day_entry["outflow"] += outflow
            total_outflow += outflow

    trend = [
        DashboardTrendPoint(
            date=day,
            inflow=_to_money(day_values["inflow"]),
            outflow=_to_money(day_values["outflow"]),
            net_amount=_to_money(day_values["net_amount"]),
            payme=_to_money(day_values["payme"]),
            click=_to_money(day_values["click"]),
            other=_to_money(day_values["other"]),
        )
        for day, day_values in daily_totals.items()
    ]

    source_breakdown = [
        DashboardSourceBreakdown(
            source=source_key,
            amount=_to_money(_to_decimal(source_totals[source_key]["amount"])),
            transaction_count=int(source_totals[source_key]["transaction_count"]),
        )
        for source_key in ("payme", "click", "other")
    ]

    return DashboardPeriodSummary(
        from_date=period_from,
        to_date=period_to,
        total_inflow=_to_money(total_inflow),
        total_outflow=_to_money(total_outflow),
        net_amount=_to_money(net_amount),
        successful_transactions=successful_transactions,
        source_breakdown=source_breakdown,
        trend=trend,
    )


def _normalize_phone(value: object) -> Optional[str]:
    if value is None:
        return None
    phone = str(value).strip()
    return phone or None


def _extract_parent_phones_from_custom_fields(custom_fields_raw: object) -> tuple[Optional[str], Optional[str], Optional[str]]:
    if not custom_fields_raw:
        return None, None, None

    payload = custom_fields_raw
    if isinstance(custom_fields_raw, str):
        try:
            payload = json.loads(custom_fields_raw)
        except Exception:
            return None, None, None

    if not isinstance(payload, dict):
        return None, None, None

    student_block = payload.get("student") if isinstance(payload.get("student"), dict) else {}
    customer_block = payload.get("buyurtmachi") if isinstance(payload.get("buyurtmachi"), dict) else {}
    primary_phone = _normalize_phone(customer_block.get("telefon"))

    father_phone = (
        _normalize_phone(student_block.get("dad_phone_number"))
        or _normalize_phone(student_block.get("father_phone"))
        or _normalize_phone(payload.get("dad_phone_number"))
        or _normalize_phone(payload.get("father_phone"))
    )
    mother_phone = (
        _normalize_phone(student_block.get("mom_phone_number"))
        or _normalize_phone(student_block.get("mother_phone"))
        or _normalize_phone(payload.get("mom_phone_number"))
        or _normalize_phone(payload.get("mother_phone"))
    )

    return father_phone, mother_phone, primary_phone


async def _get_latest_contract_terminated_map(
    db: AsyncSession,
    student_ids: list[int],
) -> dict[int, bool]:
    from app.models.domain import Contract

    if not student_ids:
        return {}

    latest_contract_subquery = (
        select(
            Contract.student_id.label("student_id"),
            Contract.status.label("contract_status"),
            func.row_number()
            .over(
                partition_by=Contract.student_id,
                order_by=Contract.created_at.desc(),
            )
            .label("row_num"),
        )
        .where(Contract.student_id.in_(student_ids))
        .subquery()
    )

    result = await db.execute(
        select(
            latest_contract_subquery.c.student_id,
            latest_contract_subquery.c.contract_status,
        ).where(latest_contract_subquery.c.row_num == 1)
    )

    return {
        row.student_id: row.contract_status == ContractStatus.TERMINATED
        for row in result.all()
    }


async def _collect_debtors_report_data(
    db: AsyncSession,
    group_id: Optional[int],
    min_debt_amount: Optional[float],
    year: Optional[int],
    month: Optional[int],
) -> list[DebtorItem]:
    from app.models.domain import Contract, Parent

    today = date.today()
    target_year = year
    if month is not None and target_year is None:
        target_year = today.year
    is_period_filter_applied = target_year is not None

    active_contract_conditions = [
        Contract.status == ContractStatus.ACTIVE,
    ]
    if group_id:
        active_contract_conditions.append(Contract.group_id == group_id)
    if not is_period_filter_applied:
        active_contract_conditions.extend(
            [
                Contract.start_date <= today,
                Contract.end_date >= today,
            ]
        )

    latest_active_contract_subquery = (
        select(
            Contract.id.label("contract_id"),
            Contract.student_id.label("student_id"),
            Contract.group_id.label("group_id"),
            Contract.contract_number.label("contract_number"),
            Contract.custom_fields.label("custom_fields"),
            Contract.start_date.label("start_date"),
            Contract.end_date.label("end_date"),
            Contract.monthly_fee.label("monthly_fee"),
            func.row_number()
            .over(
                partition_by=Contract.student_id,
                order_by=Contract.created_at.desc(),
            )
            .label("row_num"),
        )
        .where(and_(*active_contract_conditions))
        .subquery()
    )

    active_contracts_result = await db.execute(
        select(
            latest_active_contract_subquery.c.contract_id,
            latest_active_contract_subquery.c.student_id,
            latest_active_contract_subquery.c.contract_number,
            latest_active_contract_subquery.c.custom_fields,
            latest_active_contract_subquery.c.start_date,
            latest_active_contract_subquery.c.end_date,
            latest_active_contract_subquery.c.monthly_fee,
            Student.first_name,
            Student.last_name,
            Group.name.label("group_name"),
        )
        .join(Student, Student.id == latest_active_contract_subquery.c.student_id)
        .outerjoin(Group, Group.id == latest_active_contract_subquery.c.group_id)
        .where(latest_active_contract_subquery.c.row_num == 1)
    )
    active_contract_rows = active_contracts_result.all()
    if not active_contract_rows:
        return []

    contract_ids: list[int] = []
    for contract_row in active_contract_rows:
        contract_ids.append(contract_row.contract_id)

    covered_months_by_contract: dict[int, set[tuple[int, int]]] = {}
    if contract_ids:
        coverage_conditions = [
            Transaction.contract_id.in_(contract_ids),
            Transaction.status == PaymentStatus.SUCCESS,
            Transaction.payment_year.isnot(None),
        ]
        if is_period_filter_applied:
            coverage_conditions.append(Transaction.payment_year == target_year)
            if month is not None:
                coverage_conditions.append(
                    cast(Transaction.payment_months, JSONB).op("@>")(cast([month], JSONB))
                )

        coverage_result = await db.execute(
            select(
                Transaction.contract_id,
                Transaction.payment_year,
                Transaction.payment_months,
            ).where(and_(*coverage_conditions))
        )
        for row in coverage_result.all():
            if row.contract_id is None or row.payment_year is None:
                continue
            contract_covered_months = covered_months_by_contract.setdefault(row.contract_id, set())
            for month_number in row.payment_months or []:
                try:
                    month_value = int(month_number)
                except (TypeError, ValueError):
                    continue
                if 1 <= month_value <= 12:
                    contract_covered_months.add((int(row.payment_year), month_value))

    debtors_data: list[dict[str, object]] = []
    for contract_row in active_contract_rows:
        contract_start_month = contract_row.start_date.replace(day=1)
        if is_period_filter_applied:
            target_months = (
                [(target_year, month)]
                if month is not None
                else [(target_year, month_number) for month_number in range(1, 13)]
            )
            contract_end_month = contract_row.end_date.replace(day=1)
            months_in_period = [
                (target_month_year, target_month_num)
                for target_month_year, target_month_num in target_months
                if contract_start_month <= date(target_month_year, target_month_num, 1) <= contract_end_month
            ]
        else:
            period_end_month = min(today, contract_row.end_date).replace(day=1)
            months_in_period = []
            current_month = contract_start_month
            while current_month <= period_end_month:
                months_in_period.append((current_month.year, current_month.month))
                current_month = current_month + relativedelta(months=1)

        if not months_in_period:
            continue

        covered_months = covered_months_by_contract.get(contract_row.contract_id, set())
        unpaid_months = [target_month for target_month in months_in_period if target_month not in covered_months]
        unpaid_count = len(unpaid_months)
        debt = float(contract_row.monthly_fee) * unpaid_count

        if debt > 0 and (min_debt_amount is None or debt >= min_debt_amount):
            debtors_data.append(
                {
                    "student_id": contract_row.student_id,
                    "student_name": f"{contract_row.first_name} {contract_row.last_name}",
                    "contract_number": contract_row.contract_number,
                    "group_name": contract_row.group_name,
                    "debt_amount": debt,
                    "custom_fields": contract_row.custom_fields,
                    "overdue_months": unpaid_months,
                    "monthly_fee": float(contract_row.monthly_fee),
                }
            )

    if not debtors_data:
        return []

    debtor_ids = [int(item["student_id"]) for item in debtors_data]
    parent_phone_by_student: dict[int, dict[str, Optional[str]]] = {
        student_id: {"father_phone": None, "mother_phone": None}
        for student_id in debtor_ids
    }

    parent_rows = await db.execute(
        select(Parent.student_id, Parent.phone, Parent.relationship_type)
        .where(Parent.student_id.in_(debtor_ids))
        .order_by(Parent.student_id.asc(), Parent.id.asc())
    )
    father_keywords = ("father", "ota", "dad", "otets")
    mother_keywords = ("mother", "ona", "mom", "mat")
    for parent in parent_rows.all():
        if not parent.phone:
            continue
        phone_info = parent_phone_by_student.setdefault(
            int(parent.student_id), {"father_phone": None, "mother_phone": None}
        )
        relationship_type = (parent.relationship_type or "").strip().lower()

        if any(keyword in relationship_type for keyword in father_keywords):
            if not phone_info["father_phone"]:
                phone_info["father_phone"] = parent.phone
            continue
        if any(keyword in relationship_type for keyword in mother_keywords):
            if not phone_info["mother_phone"]:
                phone_info["mother_phone"] = parent.phone
            continue

        if not phone_info["father_phone"]:
            phone_info["father_phone"] = parent.phone
        elif not phone_info["mother_phone"]:
            phone_info["mother_phone"] = parent.phone

    debtor_items: list[DebtorItem] = []
    for item in debtors_data:
        student_id = int(item["student_id"])
        phone_info = parent_phone_by_student.get(student_id, {})
        father_phone = phone_info.get("father_phone")
        mother_phone = phone_info.get("mother_phone")
        primary_phone = None

        if not father_phone or not mother_phone:
            fallback_father_phone, fallback_mother_phone, fallback_primary_phone = _extract_parent_phones_from_custom_fields(
                item.get("custom_fields")
            )
            father_phone = father_phone or fallback_father_phone
            mother_phone = mother_phone or fallback_mother_phone
            primary_phone = fallback_primary_phone
        else:
            _, _, fallback_primary_phone = _extract_parent_phones_from_custom_fields(item.get("custom_fields"))
            primary_phone = fallback_primary_phone

        if not primary_phone:
            primary_phone = father_phone or mother_phone

        debtor_items.append(
            DebtorItem(
                student_id=student_id,
                student_name=str(item["student_name"]),
                contract_number=str(item["contract_number"]),
                debt_amount=float(item["debt_amount"]),
                group_name=item["group_name"],
                primary_phone=primary_phone,
                father_phone=father_phone,
                mother_phone=mother_phone,
                overdue_months=[
                    DebtorMonthItem(
                        year=year_value,
                        month=month_value,
                        amount=float(item["monthly_fee"]),
                        label=_format_month_label(year_value, month_value),
                    )
                    for year_value, month_value in item.get("overdue_months", [])
                ],
                overdue_months_count=len(item.get("overdue_months", [])),
            )
        )

    return debtor_items


async def _collect_payers_report_data(
    db: AsyncSession,
    payment_year: Optional[int],
    payment_month: Optional[int],
    group_id: Optional[int],
    min_paid_amount: Optional[float],
    from_date: Optional[date],
    to_date: Optional[date],
    offset: Optional[int] = None,
    limit: Optional[int] = None,
) -> tuple[list[PayerItem], int, int]:
    from app.models.domain import Contract
    from datetime import datetime as dt_datetime

    if payment_year is None:
        payment_year = dt_datetime.now().year

    base_conditions = [
        Transaction.status == PaymentStatus.SUCCESS,
        or_(
            Transaction.settlement_type.is_(None),
            Transaction.settlement_type == PaymentSettlementType.PAYMENT,
            Transaction.settlement_type == PaymentSettlementType.WAIVER_SPRAVKA,
        ),
        Transaction.payment_year == payment_year,
        Transaction.student_id.isnot(None),
    ]
    if payment_month is not None:
        base_conditions.append(
            cast(Transaction.payment_months, JSONB).op("@>")(cast([payment_month], JSONB))
        )

    if from_date:
        from_datetime = dt_datetime.combine(from_date, dt_datetime.min.time())
        base_conditions.append(Transaction.paid_at >= from_datetime)

    if to_date:
        to_datetime = dt_datetime.combine(to_date, dt_datetime.max.time())
        base_conditions.append(Transaction.paid_at <= to_datetime)

    aggregate_conditions = list(base_conditions)
    if group_id:
        aggregate_conditions.append(Student.group_id == group_id)

    # total_paid should reflect real money paid (Transaction.amount).
    # WAIVER_SPRAVKA keeps month coverage logic but does not add cash amount.
    total_paid_expr = func.sum(Transaction.amount)
    aggregated_base_query = (
        select(
            Student.id.label("student_id"),
            Student.first_name.label("first_name"),
            Student.last_name.label("last_name"),
            Group.name.label("group_name"),
            total_paid_expr.label("total_paid"),
        )
        .join(Student, Transaction.student_id == Student.id)
        .outerjoin(Group, Student.group_id == Group.id)
        .where(and_(*aggregate_conditions))
        .group_by(Student.id, Student.first_name, Student.last_name, Group.name)
    )

    if min_paid_amount is not None:
        aggregated_base_query = aggregated_base_query.having(total_paid_expr >= min_paid_amount)

    aggregated_subquery = aggregated_base_query.subquery()
    count_result = await db.execute(select(func.count()).select_from(aggregated_subquery))
    total = count_result.scalar() or 0

    data_query = aggregated_base_query.order_by(total_paid_expr.desc(), Student.id.asc())
    if offset is not None:
        data_query = data_query.offset(offset)
    if limit is not None:
        data_query = data_query.limit(limit)

    page_result = await db.execute(data_query)
    page_rows = page_result.all()
    if not page_rows:
        return [], total, payment_year

    page_student_ids = [row.student_id for row in page_rows]

    months_result = await db.execute(
        select(
            Transaction.student_id,
            Transaction.payment_months,
        ).where(
            and_(
                *base_conditions,
                Transaction.student_id.in_(page_student_ids),
            )
        )
    )
    months_map: dict[int, set[int]] = {}
    for row in months_result.all():
        if row.student_id not in months_map:
            months_map[row.student_id] = set()
        if row.payment_months:
            months_map[row.student_id].update(row.payment_months)

    latest_contract_subquery = (
        select(
            Transaction.student_id.label("student_id"),
            Contract.contract_number.label("contract_number"),
            func.row_number()
            .over(
                partition_by=Transaction.student_id,
                order_by=(Transaction.paid_at.desc().nullslast(), Transaction.created_at.desc()),
            )
            .label("row_num"),
        )
        .join(Contract, Contract.id == Transaction.contract_id)
        .where(
            and_(
                *base_conditions,
                Transaction.student_id.in_(page_student_ids),
                Transaction.contract_id.isnot(None),
            )
        )
        .subquery()
    )
    contract_result = await db.execute(
        select(
            latest_contract_subquery.c.student_id,
            latest_contract_subquery.c.contract_number,
        ).where(latest_contract_subquery.c.row_num == 1)
    )
    contract_map = {
        row.student_id: row.contract_number
        for row in contract_result.all()
    }

    payers = [
        PayerItem(
            student_id=row.student_id,
            student_name=f"{row.first_name} {row.last_name}",
            contract_number=contract_map.get(row.student_id),
            group_name=row.group_name,
            payment_year=payment_year,
            payment_months=sorted(list(months_map.get(row.student_id, set()))),
            total_paid=float(row.total_paid or 0.0),
        )
        for row in page_rows
    ]
    return payers, total, payment_year


@router.get("/dashboard/summary", response_model=DataResponse[DashboardSummary], dependencies=[Depends(require_permission(PERM_REPORTS_DASHBOARD_VIEW))])
async def get_dashboard_summary(db: Annotated[AsyncSession, Depends(get_db)]):
    from app.models.domain import Contract

    today = date.today()
    today_start = datetime.combine(today, datetime.min.time())
    tomorrow_start = datetime.combine(today + timedelta(days=1), datetime.min.time())
    first_day_of_month = today.replace(day=1)
    first_day_of_month_start = datetime.combine(first_day_of_month, datetime.min.time())

    today_revenue_result = await db.execute(
        select(func.sum(Transaction.amount))
        .where(
            and_(
                Transaction.status == PaymentStatus.SUCCESS,
                Transaction.paid_at >= today_start,
                Transaction.paid_at < tomorrow_start,
            )
        )
    )
    today_revenue = today_revenue_result.scalar() or 0.0

    active_students_result = await db.execute(
        select(func.count(Student.id)).where(Student.status == StudentStatus.ACTIVE)
    )
    active_students = active_students_result.scalar() or 0

    today_sessions_result = await db.execute(
        select(func.count(Session.id)).where(Session.session_date == today)
    )
    today_sessions = today_sessions_result.scalar() or 0

    # Fast debtor count: aggregate instead of N+1 per contract checks.
    active_contracts_subquery = (
        select(Contract.id)
        .where(
            and_(
                Contract.status == ContractStatus.ACTIVE,
                Contract.start_date <= today,
                Contract.end_date >= today,
            )
        )
        .subquery()
    )

    active_contracts_count_result = await db.execute(
        select(func.count()).select_from(active_contracts_subquery)
    )
    active_contracts_count = active_contracts_count_result.scalar() or 0

    covered_contracts_count_result = await db.execute(
        select(func.count(func.distinct(Transaction.contract_id))).where(
            and_(
                Transaction.contract_id.in_(select(active_contracts_subquery.c.id)),
                Transaction.status == PaymentStatus.SUCCESS,
                or_(
                    and_(
                        Transaction.payment_year == today.year,
                        cast(Transaction.payment_months, JSONB).op("@>")(cast([today.month], JSONB)),
                    ),
                    and_(
                        Transaction.payment_year.is_(None),
                        Transaction.paid_at >= first_day_of_month_start,
                    ),
                ),
            )
        )
    )
    covered_contracts_count = covered_contracts_count_result.scalar() or 0
    current_month_debtors = max(active_contracts_count - covered_contracts_count, 0)

    ninety_day_start = today - timedelta(days=89)
    transactions_result = await db.execute(
        select(
            cast(Transaction.paid_at, Date).label("paid_date"),
            Transaction.amount,
            Transaction.source,
        ).where(
            and_(
                Transaction.status == PaymentStatus.SUCCESS,
                Transaction.paid_at.isnot(None),
                Transaction.paid_at >= datetime.combine(ninety_day_start, datetime.min.time()),
                Transaction.paid_at < tomorrow_start,
            )
        )
    )
    period_rows = transactions_result.all()

    last_7_days = _build_dashboard_period_summary(
        period_from=today - timedelta(days=6),
        period_to=today,
        rows=period_rows,
    )
    last_30_days = _build_dashboard_period_summary(
        period_from=today - timedelta(days=29),
        period_to=today,
        rows=period_rows,
    )
    last_90_days = _build_dashboard_period_summary(
        period_from=ninety_day_start,
        period_to=today,
        rows=period_rows,
    )

    return DataResponse(
        data=DashboardSummary(
            today_revenue=float(today_revenue),
            active_students=active_students,
            total_debtors=current_month_debtors,
            today_sessions=today_sessions,
            last_7_days=last_7_days,
            last_30_days=last_30_days,
            last_90_days=last_90_days,
        )
    )


@router.get("/finance", response_model=DataResponse[FinanceReport], dependencies=[Depends(require_permission(PERM_REPORTS_FINANCE_VIEW))])
async def get_finance_report(
    db: Annotated[AsyncSession, Depends(get_db)],
    from_date: date = Query(...),
    to_date: date = Query(...),
    group_id: Optional[int] = Query(None, description="Filter by student group"),
    status: Optional[str] = Query(None, description="Filter by student status (active, archived, etc.)"),
):
    target_months = build_month_range(from_date, to_date)
    if not target_months:
        return DataResponse(
            data=FinanceReport(
                from_date=from_date,
                to_date=to_date,
                total_revenue=0.0,
                breakdown=[],
            )
        )

    month_set = set(target_months)
    min_year = min(year for year, _ in target_months)
    max_year = max(year for year, _ in target_months)

    conditions = [
        Transaction.status == PaymentStatus.SUCCESS,
        Transaction.student_id.isnot(None),
        Transaction.contract_id.isnot(None),
        Transaction.payment_year.isnot(None),
        Transaction.payment_year >= min_year,
        Transaction.payment_year <= max_year,
    ]
    if group_id is not None:
        conditions.append(Student.group_id == group_id)
    if status:
        conditions.append(Student.status == status)

    result = await db.execute(
        select(
            Transaction.id,
            Transaction.source,
            Transaction.amount,
            Transaction.payment_year,
            Transaction.payment_months,
        )
        .join(Student, Transaction.student_id == Student.id)
        .where(and_(*conditions))
    )

    source_total_map: dict[object, Decimal] = {}
    source_count_map: dict[object, int] = {}
    for row in result.all():
        matched_months, allocated_amount = allocate_amount_for_period(
            amount=row.amount,
            payment_year=row.payment_year,
            payment_months=row.payment_months,
            month_set=month_set,
        )
        if not matched_months:
            continue

        source_total_map[row.source] = source_total_map.get(row.source, Decimal("0")) + allocated_amount
        source_count_map[row.source] = source_count_map.get(row.source, 0) + 1

    breakdown = [
        FinanceReportItem(
            source=str(source),
            total_amount=float(total.quantize(Decimal("0.01"))),
            transaction_count=source_count_map[source],
        )
        for source, total in source_total_map.items()
    ]
    breakdown.sort(key=lambda item: item.source)

    total_revenue = float(sum(source_total_map.values(), Decimal("0")).quantize(Decimal("0.01")))

    return DataResponse(
        data=FinanceReport(
            from_date=from_date,
            to_date=to_date,
            total_revenue=total_revenue,
            breakdown=breakdown,
        )
    )


@router.get("/attendance/groups", response_model=DataResponse[list[GroupAttendanceReport]], dependencies=[Depends(require_permission(PERM_REPORTS_ATTENDANCE_VIEW))])
async def get_group_attendance_report(db: Annotated[AsyncSession, Depends(get_db)]):
    groups_result = await db.execute(select(Group))
    groups = groups_result.scalars().all()

    reports = []
    for group in groups:
        sessions_result = await db.execute(select(func.count(Session.id)).where(Session.group_id == group.id))
        total_sessions = sessions_result.scalar() or 0

        students_result = await db.execute(select(func.count(Student.id)).where(Student.group_id == group.id))
        total_students = students_result.scalar() or 0

        present_result = await db.execute(
            select(func.count(Attendance.id))
            .join(Session)
            .where(and_(Session.group_id == group.id, Attendance.status == AttendanceStatus.PRESENT))
        )
        present_count = present_result.scalar() or 0

        total_possible = total_sessions * total_students
        attendance_percentage = (present_count / total_possible * 100) if total_possible > 0 else 0.0

        reports.append(
            GroupAttendanceReport(
                group_id=group.id,
                group_name=group.name,
                total_sessions=total_sessions,
                total_students=total_students,
                attendance_percentage=attendance_percentage,
            )
        )

    return DataResponse(data=reports)


@router.get("/attendance/students/{student_id}", response_model=DataResponse[StudentAttendanceReport], dependencies=[Depends(require_permission(PERM_REPORTS_ATTENDANCE_VIEW))])
async def get_student_attendance_report(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    student_result = await db.execute(select(Student).where(Student.id == student_id))
    student = student_result.scalar_one_or_none()

    if not student:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Student not found")

    total_sessions_result = await db.execute(
        select(func.count(Session.id)).where(Session.group_id == student.group_id)
    )
    total_sessions = total_sessions_result.scalar() or 0

    present_result = await db.execute(
        select(func.count(Attendance.id))
        .where(and_(Attendance.student_id == student_id, Attendance.status == AttendanceStatus.PRESENT))
    )
    present_count = present_result.scalar() or 0

    absent_result = await db.execute(
        select(func.count(Attendance.id))
        .where(and_(Attendance.student_id == student_id, Attendance.status == AttendanceStatus.ABSENT))
    )
    absent_count = absent_result.scalar() or 0

    late_result = await db.execute(
        select(func.count(Attendance.id))
        .where(and_(Attendance.student_id == student_id, Attendance.status == AttendanceStatus.LATE))
    )
    late_count = late_result.scalar() or 0

    attendance_percentage = (present_count / total_sessions * 100) if total_sessions > 0 else 0.0

    return DataResponse(
        data=StudentAttendanceReport(
            student_id=student.id,
            student_name=f"{student.first_name} {student.last_name}",
            total_sessions=total_sessions,
            present_count=present_count,
            absent_count=absent_count,
            late_count=late_count,
            attendance_percentage=attendance_percentage,
        )
    )


@router.get("/debtors", response_model=DataResponse[list[DebtorItem]], dependencies=[Depends(require_permission(PERM_REPORTS_FINANCE_VIEW))])
async def get_debtors_report(
    db: Annotated[AsyncSession, Depends(get_db)],
    group_id: Optional[int] = None,
    min_debt_amount: Optional[float] = None,
    year: Optional[int] = Query(None, ge=2000, le=2100),
    month: Optional[int] = Query(None, ge=1, le=12),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    debtors = await _collect_debtors_report_data(
        db=db,
        group_id=group_id,
        min_debt_amount=min_debt_amount,
        year=year,
        month=month,
    )

    offset = (page - 1) * page_size
    paginated_debtors = debtors[offset : offset + page_size]

    return DataResponse(
        data=paginated_debtors,
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=len(debtors),
            total_pages=(len(debtors) + page_size - 1) // page_size,
        ),
    )


@router.get("/debtors/export", dependencies=[Depends(require_permission(PERM_REPORTS_FINANCE_VIEW))])
async def export_debtors_report(
    db: Annotated[AsyncSession, Depends(get_db)],
    group_id: Optional[int] = None,
    min_debt_amount: Optional[float] = None,
    year: Optional[int] = Query(None, ge=2000, le=2100),
    month: Optional[int] = Query(None, ge=1, le=12),
):
    today = date.today()
    resolved_year = year if year is not None else (today.year if month is not None else None)
    if resolved_year is not None and month is not None:
        period_label = f"Only {resolved_year}-{month:02d}"
    elif resolved_year is not None:
        period_label = f"All months in {resolved_year}"
    else:
        period_label = f"From contract start month to current month ({today.strftime('%Y-%m')})"

    debtors = await _collect_debtors_report_data(
        db=db,
        group_id=group_id,
        min_debt_amount=min_debt_amount,
        year=year,
        month=month,
    )

    terminated_map = await _get_latest_contract_terminated_map(
        db=db,
        student_ids=[item.student_id for item in debtors],
    )

    regular_rows: list[list[object]] = []
    terminated_rows: list[list[object]] = []
    for item in debtors:
        row = [
            item.student_id,
            item.student_name,
            item.contract_number,
            item.group_name or "",
            item.primary_phone or "",
            item.father_phone or "",
            item.mother_phone or "",
            ", ".join(month_item.label for month_item in item.overdue_months),
            float(item.debt_amount),
            period_label,
        ]
        if terminated_map.get(item.student_id, False):
            terminated_rows.append(row)
        else:
            regular_rows.append(row)

    rows = regular_rows + ([[]] if regular_rows and terminated_rows else []) + terminated_rows
    file_data = _build_excel_file(
        sheet_name="Debtors",
        headers=[
            "Student ID",
            "Student Name",
            "Contract Number",
            "Group",
            "Primary Phone",
            "Father Phone",
            "Mother Phone",
            "Debt Months",
            "Debt Amount",
            "Debt Period",
        ],
        rows=rows,
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        file_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="debtors_{timestamp}.xlsx"'},
    )


@router.get("/payers", response_model=DataResponse[list[PayerItem]], dependencies=[Depends(require_permission(PERM_REPORTS_FINANCE_VIEW))])
async def get_payers_report(
    db: Annotated[AsyncSession, Depends(get_db)],
    payment_year: Optional[int] = Query(None, description="Payment year (defaults to current year)"),
    payment_month: Optional[int] = Query(None, ge=1, le=12, description="Filter by payment month"),
    group_id: Optional[int] = None,
    min_paid_amount: Optional[float] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    """
    Get list of students who made payments with their payment details.

    This endpoint is optimized for fast performance using aggregated queries.

    Returns:
    - Student name
    - Group name
    - Contract number
    - Payment year
    - Payment months (list of months paid)
    - Total paid amount

    Filters:
    - payment_year: Optional - defaults to current year
    - payment_month: Optional - filter by specific month (1-12)
    - group_id: Optional - filter by specific group
    - min_paid_amount: Optional - minimum payment amount
    - from_date/to_date: Optional - payment date range

    Example Response:
    ```json
    {
      "data": [
        {
          "student_id": 123,
          "student_name": "Jahongir Abdullayev",
          "group_name": "Group B1",
          "contract_number": "1-2020B1",
          "payment_year": 2025,
          "payment_months": [1, 2, 3, 12],
          "total_paid": 2000000.0
        }
      ]
    }
    ```
    """
    offset = (page - 1) * page_size
    paginated_payers, total, _ = await _collect_payers_report_data(
        db=db,
        payment_year=payment_year,
        payment_month=payment_month,
        group_id=group_id,
        min_paid_amount=min_paid_amount,
        from_date=from_date,
        to_date=to_date,
        offset=offset,
        limit=page_size,
    )

    return DataResponse(
        data=paginated_payers,
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=(total + page_size - 1) // page_size,
        ),
    )


@router.get("/payers/export", dependencies=[Depends(require_permission(PERM_REPORTS_FINANCE_VIEW))])
async def export_payers_report(
    db: Annotated[AsyncSession, Depends(get_db)],
    payment_year: Optional[int] = Query(None, description="Payment year (defaults to current year)"),
    payment_month: Optional[int] = Query(None, ge=1, le=12, description="Filter by payment month"),
    group_id: Optional[int] = None,
    min_paid_amount: Optional[float] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
):
    payers, _, resolved_payment_year = await _collect_payers_report_data(
        db=db,
        payment_year=payment_year,
        payment_month=payment_month,
        group_id=group_id,
        min_paid_amount=min_paid_amount,
        from_date=from_date,
        to_date=to_date,
    )

    terminated_map = await _get_latest_contract_terminated_map(
        db=db,
        student_ids=[item.student_id for item in payers],
    )

    regular_rows: list[list[object]] = []
    terminated_rows: list[list[object]] = []
    for item in payers:
        row = [
            item.student_id,
            item.student_name,
            item.contract_number or "",
            item.group_name or "",
            item.payment_year,
            ", ".join(str(month_number) for month_number in item.payment_months),
            float(item.total_paid),
        ]
        if terminated_map.get(item.student_id, False):
            terminated_rows.append(row)
        else:
            regular_rows.append(row)

    rows = regular_rows + ([[]] if regular_rows and terminated_rows else []) + terminated_rows
    file_data = _build_excel_file(
        sheet_name="Payers",
        headers=["Student ID", "Student Name", "Contract Number", "Group", "Payment Year", "Payment Months", "Total Paid"],
        rows=rows,
    )

    if payment_month is not None:
        file_suffix = f"{resolved_payment_year}_{payment_month:02d}"
    else:
        file_suffix = str(resolved_payment_year)

    return StreamingResponse(
        file_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="payers_{file_suffix}.xlsx"'},
    )

