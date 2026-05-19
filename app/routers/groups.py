from typing import Annotated
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_
from collections import defaultdict
from app.core.db import get_db
from app.core.permissions import PERM_GROUPS_VIEW, PERM_GROUPS_EDIT
from app.models.domain import Group, Student, WaitingList
from app.schemas.group import GroupRead, GroupCreate, GroupUpdate, GroupCapacityInfo, GroupCapacityByYear, GroupsByYear, GroupedByYearResponse, GroupStatistics, GroupStatisticsByBirthYear
from app.schemas.student import StudentRead
from app.schemas.common import DataResponse, PaginationMeta
from app.deps import require_permission
from app.models.enums import GroupStatus, StudentStatus

router = APIRouter(prefix="/groups", tags=["Groups"])


@router.get("", response_model=DataResponse[list[GroupRead]], dependencies=[Depends(require_permission(PERM_GROUPS_VIEW))])
async def get_groups(
    db: Annotated[AsyncSession, Depends(get_db)],
    archive_year: int | None = Query(None, description="Filter by archive year (NULL for groups not yet archived)"),
    birth_year: int | None = Query(None, description="Filter by birth year (e.g., 2020, 2019)"),
    status: GroupStatus | None = Query(None, description="Filter by status (ACTIVE, ARCHIVED, DELETED)"),
    include_archived: bool = Query(False, description="Include archived groups (ignored if status is specified)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    """
    Get all groups with optional filters.

    Default behavior (no filters):
    - Shows all ACTIVE groups regardless of archive_year

    Filters:
    - archive_year: Filter by specific year, or use NULL to find groups not yet archived
    - birth_year: Filter by student birth year (e.g., 2020 shows all groups for students born in 2020)
    - status: Filter by specific status (overrides include_archived)
    - include_archived: Include all statuses if True (ignored if status is specified)
    """
    query = select(Group)

    # Filter by archive_year if specified
    if archive_year is not None:
        query = query.where(Group.archive_year == archive_year)

    # Filter by birth_year if specified
    if birth_year is not None:
        query = query.where(Group.birth_year == birth_year)

    # Filter by status if specified (takes priority)
    if status is not None:
        query = query.where(Group.status == status)
    # Otherwise, default to ACTIVE only unless include_archived is True
    elif not include_archived:
        query = query.where(Group.status == GroupStatus.ACTIVE)

    offset = (page - 1) * page_size
    result = await db.execute(query.offset(offset).limit(page_size))
    groups = result.scalars().all()

    count_query = select(func.count(Group.id))
    if archive_year is not None:
        count_query = count_query.where(Group.archive_year == archive_year)
    if birth_year is not None:
        count_query = count_query.where(Group.birth_year == birth_year)
    if status is not None:
        count_query = count_query.where(Group.status == status)
    elif not include_archived:
        count_query = count_query.where(Group.status == GroupStatus.ACTIVE)

    count_result = await db.execute(count_query)
    total = count_result.scalar()

    # Enrich groups with capacity usage and waiting list counts
    groups_data = []
    for group in groups:
        group_dict = GroupRead.model_validate(group).model_dump()

        # Count active students in this group
        student_count_result = await db.execute(
            select(func.count(Student.id)).where(
                and_(
                    Student.group_id == group.id,
                    Student.status == StudentStatus.ACTIVE,
                    Student.archive_year == group.archive_year,
                )
            )
        )
        group_dict['active_students_count'] = student_count_result.scalar() or 0

        # Count waiting list entries
        waiting_count_result = await db.execute(
            select(func.count(WaitingList.id)).where(WaitingList.group_id == group.id)
        )
        group_dict['waiting_list_count'] = waiting_count_result.scalar() or 0

        groups_data.append(group_dict)

    return DataResponse(
        data=groups_data,
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=(total + page_size - 1) // page_size,
        ),
    )


@router.get("/grouped-by-year", response_model=GroupedByYearResponse, dependencies=[Depends(require_permission(PERM_GROUPS_VIEW))])
async def get_groups_grouped_by_year(
    db: Annotated[AsyncSession, Depends(get_db)],
    archive_year: int | None = Query(None, description="Filter by archive year"),
    status: GroupStatus | None = Query(None, description="Filter by status (ACTIVE, ARCHIVED, DELETED)"),
    include_archived: bool = Query(False, description="Include archived groups"),
):
    """
    Get all groups organized by birth year.

    Returns groups grouped by birth year with statistics for each year.
    Useful for seeing the distribution of groups across different age categories.

    Example Response:
    ```json
    {
      "data": [
        {
          "birth_year": 2020,
          "groups": [
            {"id": 1, "name": "Group B1", "identifier": "B1", ...},
            {"id": 2, "name": "Group B2", "identifier": "B2", ...}
          ],
          "total_groups": 2
        },
        {
          "birth_year": 2019,
          "groups": [
            {"id": 3, "name": "Group C1", "identifier": "C1", ...}
          ],
          "total_groups": 1
        }
      ],
      "total_birth_years": 2
    }
    ```
    """
    query = select(Group)

    # Apply filters
    if archive_year is not None:
        query = query.where(Group.archive_year == archive_year)

    if status is not None:
        query = query.where(Group.status == status)
    elif not include_archived:
        query = query.where(Group.status == GroupStatus.ACTIVE)

    # Order by birth_year descending (newest first)
    query = query.order_by(Group.birth_year.desc())

    result = await db.execute(query)
    groups = result.scalars().all()

    # Group by birth year
    groups_by_year = defaultdict(list)
    for group in groups:
        groups_by_year[group.birth_year].append(group)

    # Convert to response format
    grouped_data = []
    for birth_year in sorted(groups_by_year.keys(), reverse=True):
        year_groups = groups_by_year[birth_year]

        # Enrich each group with student counts
        enriched_groups = []
        for group in year_groups:
            group_dict = GroupRead.model_validate(group).model_dump()

            # Count active students
            student_count_result = await db.execute(
                select(func.count(Student.id)).where(
                    and_(
                        Student.group_id == group.id,
                        Student.status == StudentStatus.ACTIVE,
                        Student.archive_year == group.archive_year,
                    )
                )
            )
            group_dict['active_students_count'] = student_count_result.scalar() or 0

            # Count waiting list
            waiting_count_result = await db.execute(
                select(func.count(WaitingList.id)).where(WaitingList.group_id == group.id)
            )
            group_dict['waiting_list_count'] = waiting_count_result.scalar() or 0

            enriched_groups.append(GroupRead(**group_dict))

        grouped_data.append(GroupsByYear(
            birth_year=birth_year,
            groups=enriched_groups,
            total_groups=len(year_groups)
        ))

    return GroupedByYearResponse(
        data=grouped_data,
        total_birth_years=len(groups_by_year)
    )


@router.post("", response_model=DataResponse[GroupRead], dependencies=[Depends(require_permission(PERM_GROUPS_EDIT))])
async def create_group(
    data: GroupCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Create a new group with current year as archive year.

    Identifier must be unique (enforced by database constraint).
    Multiple groups can have the same birth year as long as identifiers are different.
    """
    # Check if identifier already exists (excluding DELETED groups)
    existing_identifier = await db.execute(
        select(Group).where(
            and_(
                Group.identifier == data.identifier,
                Group.birth_year == data.birth_year,
                Group.status != GroupStatus.DELETED
            )
        )
    )
    if existing_identifier.scalars().first():
        raise HTTPException(
            status_code=400,
            detail=f"Identifier '{data.identifier}' already exists. Please use a unique identifier."
        )

    if data.coach_id:
        from app.models.auth import User
        coach_result = await db.execute(select(User).where(User.id == data.coach_id))
        if not coach_result.scalar_one_or_none():
            raise HTTPException(status_code=404, detail=f"Coach with ID {data.coach_id} not found")

    from datetime import datetime
    current_year = datetime.now().year

    group_data = data.model_dump()
    group_data['archive_year'] = current_year  # Auto-set to current year (2025, 2026, etc.)
    group = Group(**group_data)
    db.add(group)

    try:
        await db.commit()
        await db.refresh(group)
    except Exception as e:
        await db.rollback()
        # Check if it's a unique constraint violation
        error_msg = str(e).lower()
        if "duplicate" in error_msg or "unique" in error_msg:
            if "identifier" in error_msg:
                raise HTTPException(
                    status_code=400,
                    detail=f"Identifier '{data.identifier}' already exists. Please use a unique identifier."
                )
        # Re-raise if it's a different error
        raise HTTPException(
            status_code=500,
            detail=f"Failed to create group: {str(e)}"
        )

    return DataResponse(data=GroupRead.model_validate(group))


@router.get("/statistics", response_model=DataResponse[GroupStatistics], dependencies=[Depends(require_permission(PERM_GROUPS_VIEW))])
async def get_groups_statistics(
    db: Annotated[AsyncSession, Depends(get_db)],
    archive_year: int | None = Query(None, description="Filter by archive year (defaults to current year)"),
    status: GroupStatus | None = Query(None, description="Filter by status (default: ACTIVE)"),
):
    """
    Get overall statistics for all groups.

    Returns:
    - Total groups count
    - Total capacity (sum of all group capacities)
    - Total used spots (active students)
    - Total available spots
    - Number of filled groups (groups where capacity == active students)
    - Statistics grouped by birth year

    Example Response:
    ```json
    {
      "data": {
        "total_groups": 15,
        "total_capacity": 375,
        "total_used": 320,
        "total_available": 55,
        "filled_groups_count": 3,
        "by_birth_year": [
          {
            "birth_year": 2020,
            "total_groups": 5,
            "total_capacity": 125,
            "total_used": 110,
            "total_available": 15
          },
          ...
        ]
      }
    }
    ```
    """
    from datetime import datetime
    if archive_year is None:
        archive_year = datetime.now().year

    # Default to ACTIVE status if not specified
    if status is None:
        status = GroupStatus.ACTIVE

    # Get all groups matching filters
    query = select(Group).where(Group.archive_year == archive_year)
    if status:
        query = query.where(Group.status == status)

    result = await db.execute(query)
    groups = result.scalars().all()

    # Calculate overall statistics
    total_groups = len(groups)
    total_capacity = sum(g.capacity for g in groups)

    # Get active students count for each group
    group_stats = {}
    for group in groups:
        students_result = await db.execute(
            select(func.count(Student.id)).where(
                and_(
                    Student.group_id == group.id,
                    Student.archive_year == archive_year,
                    Student.status == StudentStatus.ACTIVE
                )
            )
        )
        active_students_count = students_result.scalar() or 0

        group_stats[group.id] = {
            'birth_year': group.birth_year,
            'capacity': group.capacity,
            'used': active_students_count,
            'available': group.capacity - active_students_count
        }

    # Calculate total used and available
    total_used = sum(stats['used'] for stats in group_stats.values())
    total_available = total_capacity - total_used

    # Count filled groups (where capacity == used)
    filled_groups_count = sum(1 for stats in group_stats.values() if stats['capacity'] == stats['used'])

    # Group statistics by birth year
    by_birth_year = defaultdict(lambda: {'total_groups': 0, 'total_capacity': 0, 'total_used': 0, 'total_available': 0})

    for group_id, stats in group_stats.items():
        birth_year = stats['birth_year']
        by_birth_year[birth_year]['total_groups'] += 1
        by_birth_year[birth_year]['total_capacity'] += stats['capacity']
        by_birth_year[birth_year]['total_used'] += stats['used']
        by_birth_year[birth_year]['total_available'] += stats['available']

    # Convert to list and sort by birth year
    by_birth_year_list = [
        GroupStatisticsByBirthYear(
            birth_year=birth_year,
            total_groups=data['total_groups'],
            total_capacity=data['total_capacity'],
            total_used=data['total_used'],
            total_available=data['total_available']
        )
        for birth_year, data in sorted(by_birth_year.items(), reverse=True)
    ]

    return DataResponse(data=GroupStatistics(
        total_groups=total_groups,
        total_capacity=total_capacity,
        total_used=total_used,
        total_available=total_available,
        filled_groups_count=filled_groups_count,
        by_birth_year=by_birth_year_list
    ))


@router.get("/{group_id}", response_model=DataResponse[GroupRead], dependencies=[Depends(require_permission(PERM_GROUPS_VIEW))])
async def get_group(
    group_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(Group).where(Group.id == group_id))
    group = result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    return DataResponse(data=GroupRead.model_validate(group))


@router.patch("/{group_id}", response_model=DataResponse[GroupRead], dependencies=[Depends(require_permission(PERM_GROUPS_EDIT))])
async def update_group(
    group_id: int,
    data: GroupUpdate,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Update group details.

    Note: Identifier cannot be changed after creation for data integrity.
    Multiple groups can have the same birth year as long as identifiers are different.
    """
    result = await db.execute(select(Group).where(Group.id == group_id))
    group = result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    update_data = data.model_dump(exclude_unset=True)

    # Prevent identifier from being changed (only if it's actually different)
    if "identifier" in update_data and update_data["identifier"] != group.identifier:
        raise HTTPException(
            status_code=400,
            detail="Identifier cannot be changed after group creation. Please create a new group if you need a different identifier."
        )

    # Prevent birth_year from being changed
    if "birth_year" in update_data and update_data["birth_year"] != group.birth_year:
        raise HTTPException(
            status_code=400,
            detail="Birth year cannot be changed after group creation."
        )

    # Remove identifier and birth_year from update_data to prevent any accidental changes
    update_data.pop("identifier", None)
    update_data.pop("birth_year", None)

    if "coach_id" in update_data and update_data["coach_id"] is not None:
        from app.models.auth import User
        coach_result = await db.execute(select(User).where(User.id == update_data["coach_id"]))
        if not coach_result.scalar_one_or_none():
            raise HTTPException(status_code=404, detail=f"Coach with ID {update_data['coach_id']} not found")

    for field, value in update_data.items():
        setattr(group, field, value)

    try:
        await db.commit()
        await db.refresh(group)
    except Exception as e:
        await db.rollback()
        # Check if it's a unique constraint violation
        error_msg = str(e).lower()
        if "duplicate" in error_msg or "unique" in error_msg:
            if "identifier" in error_msg:
                identifier_value = update_data.get("identifier", group.identifier)
                raise HTTPException(
                    status_code=400,
                    detail=f"Identifier '{identifier_value}' already exists. Please use a unique identifier."
                )
        # Re-raise if it's a different error
        raise HTTPException(
            status_code=500,
            detail=f"Failed to update group: {str(e)}"
        )

    return DataResponse(data=GroupRead.model_validate(group))


@router.get("/{group_id}/students", response_model=DataResponse[list[StudentRead]], dependencies=[Depends(require_permission(PERM_GROUPS_VIEW))])
async def get_group_students(
    group_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Get only active students in a specific group."""
    result = await db.execute(
        select(Student)
        .where(
            Student.group_id == group_id,
            Student.status.notin_([StudentStatus.DELETED, StudentStatus.TERMINATED]),
        )
        .order_by(Student.last_name.asc(), Student.first_name.asc())
    )
    students = result.scalars().all()
    return DataResponse(data=[StudentRead.model_validate(s) for s in students])


@router.delete("/{group_id}", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_GROUPS_EDIT))])
async def delete_group(
    group_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Soft delete a group by setting its status to DELETED.
    The group is not actually removed from the database.
    """
    result = await db.execute(select(Group).where(Group.id == group_id))
    group = result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    # Soft delete: set status to DELETED instead of actually deleting
    group.status = GroupStatus.DELETED
    await db.commit()

    return DataResponse(data={"message": "Group deleted successfully"})


@router.post("/bulk-delete", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_GROUPS_EDIT))])
async def bulk_delete_groups(
    group_ids: list[int],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Soft delete multiple groups by setting their status to DELETED.
    Groups are not actually removed from the database.
    """
    if not group_ids:
        raise HTTPException(status_code=400, detail="No group IDs provided")

    deleted_count = 0
    errors = []

    for group_id in group_ids:
        try:
            result = await db.execute(select(Group).where(Group.id == group_id))
            group = result.scalar_one_or_none()

            if not group:
                errors.append({"group_id": group_id, "error": "Group not found"})
                continue

            # Soft delete: set status to DELETED instead of actually deleting
            group.status = GroupStatus.DELETED
            deleted_count += 1
        except Exception as e:
            errors.append({"group_id": group_id, "error": str(e)})

    await db.commit()

    return DataResponse(data={
        "message": f"Deleted {deleted_count} group(s)",
        "deleted_count": deleted_count,
        "total_requested": len(group_ids),
        "errors": errors if errors else None
    })


@router.get("/{group_id}/capacity", response_model=DataResponse[GroupCapacityInfo], dependencies=[Depends(require_permission(PERM_GROUPS_VIEW))])
async def get_group_capacity(
    group_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    archive_year: int | None = Query(None, description="Filter by archive year (defaults to current year)"),
):
    """
    Get detailed capacity information for a group.

    Returns:
    - Group capacity and current usage
    - Breakdown by birth year (how many students from each birth year)
    - Available slots
    - Waiting list count

    Useful for:
    - Checking if group has space for new students
    - Seeing distribution of students by birth year
    - Managing group capacity
    """
    from datetime import datetime
    if archive_year is None:
        archive_year = datetime.now().year

    # Get group
    group_result = await db.execute(select(Group).where(Group.id == group_id))
    group = group_result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    # Get active students for this group for the archive year
    students_result = await db.execute(
        select(Student).where(
            and_(
                Student.group_id == group_id,
                Student.archive_year == archive_year,
                Student.status == StudentStatus.ACTIVE,
            )
        )
    )
    students = students_result.scalars().all()
    active_students_count = len(students)

    # Group students by birth year
    by_year = defaultdict(lambda: {"used": 0, "available": 0})

    by_year[group.birth_year]["used"] = active_students_count

    # Calculate available slots for each year
    for year_str in by_year:
        by_year[year_str]["available"] = group.capacity - by_year[year_str]["used"]

    # Get waiting list count
    waiting_result = await db.execute(
        select(func.count(WaitingList.id)).where(WaitingList.group_id == group_id)
    )
    waiting_count = waiting_result.scalar() or 0

    # Convert by_year to the schema format
    by_year_dict = {
        str(year): GroupCapacityByYear(used=data["used"], available=data["available"])
        for year, data in by_year.items()
    }

    return DataResponse(data=GroupCapacityInfo(
        group_id=group_id,
        group_name=group.name,
        capacity=group.capacity,
        active_students=active_students_count,
        available_slots=group.capacity - active_students_count,
        waiting_list_count=waiting_count,
        by_birth_year=by_year_dict
    ))


@router.get("/{group_id}/export-students", dependencies=[Depends(require_permission(PERM_GROUPS_VIEW))])
async def export_group_students(
        group_id: int,
        db: Annotated[AsyncSession, Depends(get_db)],
):

    from fastapi.responses import FileResponse
    import os
    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    # Verify group exists
    group_result = await db.execute(select(Group).where(Group.id == group_id))
    group = group_result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    result = await db.execute(
        select(Student)
        .where(
            Student.group_id == group_id,
            Student.status == StudentStatus.ACTIVE,
        )
        .order_by(Student.last_name.asc(), Student.first_name.asc())
    )
    students = result.scalars().all()

    if not students:
        raise HTTPException(
            status_code=404,
            detail=f"No active students found for group {group.name}"
        )

    # Create Excel workbook
    wb = Workbook()
    sheet = wb.active
    sheet.title = f"Group {group.identifier}"

    # Set column widths
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 18
    sheet.column_dimensions['F'].width = 18

    # Header row styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = [
        'Student ID',
        'First Name',
        'Last Name',
        'PNFL',
        'Phone',
        'Status',
    ]

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    for row_num, student in enumerate(students, 2):
        sheet.cell(row=row_num, column=1).value = student.id
        sheet.cell(row=row_num, column=2).value = student.first_name
        sheet.cell(row=row_num, column=3).value = student.last_name
        sheet.cell(row=row_num, column=4).value = student.pnfl
        sheet.cell(row=row_num, column=5).value = student.phone or ""
        sheet.cell(row=row_num, column=6).value = student.status.value

    # Save to temporary file
    temp_dir = tempfile.gettempdir()
    filename = f"group_{group.identifier}_students.xlsx"
    filepath = os.path.join(temp_dir, filename)

    wb.save(filepath)

    return FileResponse(
        path=filepath,
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
    )
