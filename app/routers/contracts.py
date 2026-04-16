from typing import Annotated, Optional, List
from datetime import datetime, date
import json
import uuid
import os
import re
from urllib.parse import quote
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.core.db import get_db
from app.core.permissions import PERM_CONTRACTS_VIEW, PERM_CONTRACTS_EDIT
from app.core.s3 import upload_image_to_s3
from app.models.domain import Contract, Student, Group, WaitingList
from app.schemas.contract import (
    ContractRead, ContractUpdate, ContractTerminate,
    ContractCreateWithDocuments, ContractCreatedResponse,
    ContractNumberInfo, ContractCustomFields, ContractReadWithStudentName,
    ContractDatesUpdate, TerminatedStudentRead
)
from app.schemas.common import DataResponse, PaginationMeta
from app.deps import require_permission, CurrentUser
from app.models.auth import User
from app.models.enums import ContractStatus, StudentStatus
from app.services.contract_allocation import (
    get_available_contract_numbers,
    is_group_full,
    ContractNumberAllocationError
)
# from app.utils.pdf_generator import ContractGenerator, convert_dates

router = APIRouter(prefix="/contracts", tags=["Contracts"])


async def _generate_unique_terminated_contract_number(
    db: AsyncSession,
    base_contract_number: str,
    termination_dt: datetime,
    current_contract_id: int,
) -> str:
    """
    Generate a unique terminated contract number in the format:
    - {base}-T-yyyymmdd
    - {base}-T-yyyymmdd-2 (if needed)
    """
    date_suffix = termination_dt.date().strftime("%Y%m%d")
    candidate_base = f"{base_contract_number}-T-{date_suffix}"
    candidate = candidate_base
    counter = 2

    while True:
        existing = await db.execute(
            select(Contract.id).where(
                and_(
                    Contract.contract_number == candidate,
                    Contract.id != current_contract_id,
                )
            )
        )
        if existing.scalar_one_or_none() is None:
            return candidate
        candidate = f"{candidate_base}-{counter}"
        counter += 1


def _extract_original_contract_number(contract_number: str) -> str:
    marker = "-T-"
    if marker in contract_number:
        return contract_number.split(marker, 1)[0]
    return contract_number


def _resolve_terminated_by_full_name(terminated_by: User | None) -> str | None:
    if not terminated_by:
        return None
    full_name = getattr(terminated_by, "full_name", None)
    if full_name:
        return full_name
    first_name = getattr(terminated_by, "first_name", "")
    last_name = getattr(terminated_by, "last_name", "")
    fallback_name = f"{first_name} {last_name}".strip()
    return fallback_name or None


def _build_terminated_contract_filters(
    archive_year: int,
    group_id: int | None,
    search: str | None,
    terminated_from: date | None,
    terminated_to: date | None,
) -> list:
    filters = [
        Contract.archive_year == archive_year,
        Contract.status == ContractStatus.TERMINATED,
    ]

    if group_id is not None:
        filters.append(Contract.group_id == group_id)

    if terminated_from is not None:
        filters.append(func.date(Contract.terminated_at) >= terminated_from)

    if terminated_to is not None:
        filters.append(func.date(Contract.terminated_at) <= terminated_to)

    if search:
        search_value = f"%{search}%"
        filters.append(
            or_(
                Contract.contract_number.ilike(search_value),
                Student.first_name.ilike(search_value),
                Student.last_name.ilike(search_value),
                Student.phone.ilike(search_value),
            )
        )

    return filters


@router.get("", response_model=DataResponse[list[ContractRead]], dependencies=[Depends(require_permission(PERM_CONTRACTS_VIEW))])
async def get_contracts(
    db: Annotated[AsyncSession, Depends(get_db)],
    archive_year: int | None = Query(None, description="Filter by archive year (defaults to current year)"),
    include_archived: bool = Query(False, description="Include archived contracts (default: excludes ARCHIVED)"),
    status: Optional[str] = None,
    student_id: Optional[int] = None,
    group_id: Optional[int] = None,
    contract_number: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=2000),
):
    """
    Get all contracts with optional filters.

    Default behavior:
    - Shows current year's contracts only
    - Shows only ACTIVE contracts
    - Set include_archived=true to also show ARCHIVED contracts
    - TERMINATED contracts are available via /contracts/terminated-students endpoint
    """
    # Default to current year if not specified
    if archive_year is None:
        archive_year = datetime.now().year

    query = select(Contract).where(Contract.archive_year == archive_year)

    # Apply status filtering based on parameters
    if status:
        # If specific status requested, use it
        query = query.where(Contract.status == status)
    elif not include_archived:
        # Default: show only ACTIVE contracts
        query = query.where(Contract.status == ContractStatus.ACTIVE)
    else:
        # include_archived=true: show all non-DELETED statuses
        query = query.where(Contract.status.in_([
            ContractStatus.ACTIVE,
            ContractStatus.EXPIRED,
            ContractStatus.ARCHIVED,
            ContractStatus.TERMINATED,
        ]))
    if student_id:
        query = query.where(Contract.student_id == student_id)
    if group_id:
        query = query.where(Contract.group_id == group_id)
    if contract_number:
        query = query.where(Contract.contract_number.ilike(f"%{contract_number}%"))

    offset = (page - 1) * page_size
    # Order by most recent first
    query = query.order_by(Contract.created_at.desc())
    result = await db.execute(query.offset(offset).limit(page_size))
    contracts = result.scalars().all()

    count_query = select(func.count(Contract.id)).where(Contract.archive_year == archive_year)

    # Apply same status filtering to count
    if status:
        count_query = count_query.where(Contract.status == status)
    elif not include_archived:
        count_query = count_query.where(Contract.status == ContractStatus.ACTIVE)
    else:
        count_query = count_query.where(Contract.status.in_([
            ContractStatus.ACTIVE,
            ContractStatus.EXPIRED,
            ContractStatus.ARCHIVED,
            ContractStatus.TERMINATED,
        ]))
    if student_id:
        count_query = count_query.where(Contract.student_id == student_id)
    if group_id:
        count_query = count_query.where(Contract.group_id == group_id)
    if contract_number:
        count_query = count_query.where(Contract.contract_number.ilike(f"%{contract_number}%"))

    count_result = await db.execute(count_query)
    total = count_result.scalar()

    return DataResponse(
        data=[ContractRead.model_validate(c) for c in contracts],
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=(total + page_size - 1) // page_size,
        ),
    )


@router.get("/withname", response_model=DataResponse[list[ContractReadWithStudentName]], dependencies=[Depends(require_permission(PERM_CONTRACTS_VIEW))])
async def get_contracts_with_student_name(
    db: Annotated[AsyncSession, Depends(get_db)],
    archive_year: int | None = Query(None, description="Filter by archive year (defaults to current year)"),
    include_archived: bool = Query(False, description="Include archived contracts (default: excludes ARCHIVED)"),
    status: Optional[str] = None,
    student_id: Optional[int] = None,
    group_id: Optional[int] = None,
    contract_number: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=2000),
):
    """
    Get all contracts with student full name instead of student_id.

    Same functionality as GET /contracts but returns student's full name (first_name + last_name)
    instead of student_id.

    Default behavior:
    - Shows current year's contracts only
    - Shows only ACTIVE contracts
    - Set include_archived=true to also show ARCHIVED contracts
    - TERMINATED contracts are available via /contracts/terminated-students endpoint
    """
    # Default to current year if not specified
    if archive_year is None:
        archive_year = datetime.now().year

    # Join with Student table to get student names
    query = select(Contract, Student).join(
        Student, Contract.student_id == Student.id
    ).where(Contract.archive_year == archive_year)

    # Apply status filtering based on parameters
    if status:
        # If specific status requested, use it
        query = query.where(Contract.status == status)
    elif not include_archived:
        # Default: show only ACTIVE contracts
        query = query.where(Contract.status == ContractStatus.ACTIVE)
    else:
        # include_archived=true: show all non-DELETED statuses
        query = query.where(Contract.status.in_([
            ContractStatus.ACTIVE,
            ContractStatus.EXPIRED,
            ContractStatus.ARCHIVED,
            ContractStatus.TERMINATED,
        ]))
    if student_id:
        query = query.where(Contract.student_id == student_id)
    if group_id:
        query = query.where(Contract.group_id == group_id)
    if contract_number:
        query = query.where(Contract.contract_number.ilike(f"%{contract_number}%"))

    offset = (page - 1) * page_size
    # Order by most recent first
    query = query.order_by(Contract.created_at.desc())
    result = await db.execute(query.offset(offset).limit(page_size))
    rows = result.all()

    # Build response with student full name
    contracts_with_names = []
    for contract, student in rows:
        contract_dict = {
            "id": contract.id,
            "contract_number": contract.contract_number,
            "start_date": contract.start_date,
            "end_date": contract.end_date,
            "status": contract.status,
            "student_full_name": f"{student.first_name} {student.last_name}",
            "group_id": contract.group_id,
            "birth_year": contract.birth_year,
            "sequence_number": contract.sequence_number,
            "passport_copy_url": contract.passport_copy_url,
            "form_086_url": contract.form_086_url,
            "heart_checkup_url": contract.heart_checkup_url,
            "birth_certificate_url": contract.birth_certificate_url,
            "contract_images_urls": contract.contract_images_urls,
            "final_pdf_url": contract.final_pdf_url,
            "custom_fields": contract.custom_fields,
            "terminated_at": contract.terminated_at,
            "terminated_by_user_id": contract.terminated_by_user_id,
            "termination_reason": contract.termination_reason,
            "terminated_by": contract.terminated_by,
            "created_at": contract.created_at,
        }
        contracts_with_names.append(ContractReadWithStudentName(**contract_dict))

    count_query = select(func.count(Contract.id)).where(Contract.archive_year == archive_year)

    # Apply same status filtering to count
    if status:
        count_query = count_query.where(Contract.status == status)
    elif not include_archived:
        count_query = count_query.where(Contract.status == ContractStatus.ACTIVE)
    else:
        count_query = count_query.where(Contract.status.in_([
            ContractStatus.ACTIVE,
            ContractStatus.EXPIRED,
            ContractStatus.ARCHIVED,
            ContractStatus.TERMINATED,
        ]))
    if student_id:
        count_query = count_query.where(Contract.student_id == student_id)
    if group_id:
        count_query = count_query.where(Contract.group_id == group_id)
    if contract_number:
        count_query = count_query.where(Contract.contract_number.ilike(f"%{contract_number}%"))

    count_result = await db.execute(count_query)
    total = count_result.scalar()

    return DataResponse(
        data=contracts_with_names,
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=(total + page_size - 1) // page_size,
        ),
    )


@router.get(
    "/terminated-students",
    response_model=DataResponse[list[TerminatedStudentRead]],
    dependencies=[Depends(require_permission(PERM_CONTRACTS_VIEW))],
)
async def get_terminated_students(
    db: Annotated[AsyncSession, Depends(get_db)],
    archive_year: int | None = Query(None, description="Filter by archive year (defaults to current year)"),
    group_id: Optional[int] = Query(None, description="Filter by group"),
    search: Optional[str] = Query(None, description="Search by contract number or student name/phone"),
    terminated_from: Optional[date] = Query(None, description="Termination start date (YYYY-MM-DD)"),
    terminated_to: Optional[date] = Query(None, description="Termination end date (YYYY-MM-DD)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=2000),
):
    """
    Get terminated students with full contract and student information.
    """
    if archive_year is None:
        archive_year = datetime.now().year

    if terminated_from and terminated_to and terminated_from > terminated_to:
        raise HTTPException(status_code=400, detail="terminated_from must be before or equal to terminated_to")

    filters = _build_terminated_contract_filters(
        archive_year=archive_year,
        group_id=group_id,
        search=search,
        terminated_from=terminated_from,
        terminated_to=terminated_to,
    )

    query = (
        select(Contract)
        .join(Student, Contract.student_id == Student.id)
        .options(
            selectinload(Contract.student),
            selectinload(Contract.group),
            selectinload(Contract.terminated_by),
        )
        .where(and_(*filters))
        .order_by(Contract.terminated_at.desc().nullslast(), Contract.created_at.desc())
    )

    offset = (page - 1) * page_size
    result = await db.execute(query.offset(offset).limit(page_size))
    contracts = result.scalars().all()

    count_result = await db.execute(
        select(func.count(Contract.id))
        .join(Student, Contract.student_id == Student.id)
        .where(and_(*filters))
    )
    total = count_result.scalar() or 0

    data: list[TerminatedStudentRead] = []
    for contract in contracts:
        student = contract.student
        group = contract.group
        terminated_by = contract.terminated_by
        terminated_by_full_name = _resolve_terminated_by_full_name(terminated_by)

        data.append(
            TerminatedStudentRead(
                contract_id=contract.id,
                contract_number=contract.contract_number,
                original_contract_number=_extract_original_contract_number(contract.contract_number),
                start_date=contract.start_date,
                end_date=contract.end_date,
                contract_status=contract.status,
                student_id=student.id if student else contract.student_id,
                student_first_name=student.first_name if student else "",
                student_last_name=student.last_name if student else "",
                student_height=student.height if student else 0,
                student_weight=student.weight if student else 0,
                student_pnfl=student.pnfl if student else "",
                student_phone=student.phone if student else None,
                student_group_id=student.group_id if student else contract.group_id,
                student_group_name=group.name if group else None,
                student_group_identifier=group.identifier if group else None,
                terminated_at=contract.terminated_at,
                termination_reason=contract.termination_reason,
                terminated_by_user_id=contract.terminated_by_user_id,
                terminated_by_full_name=terminated_by_full_name,
            )
        )

    return DataResponse(
        data=data,
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=(total + page_size - 1) // page_size,
        ),
    )


'''
@router.get(
    "/terminated-unpaid-report",
    response_model=DataResponse[list[TerminatedUnpaidReportRow]],
    dependencies=[Depends(require_permission(PERM_CONTRACTS_VIEW))],
)
async def get_terminated_unpaid_report(
    db: Annotated[AsyncSession, Depends(get_db)],
    archive_year: int | None = Query(None, description="Filter by archive year (defaults to current year)"),
    group_id: Optional[int] = Query(None, description="Filter by group"),
    search: Optional[str] = Query(None, description="Search by contract number or student name/phone"),
    terminated_from: Optional[date] = Query(None, description="Termination start date (YYYY-MM-DD)"),
    terminated_to: Optional[date] = Query(None, description="Termination end date (YYYY-MM-DD)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=2000),
):
    """
    Terminated contracts report with paid/unpaid months and debt details.
    """
    if archive_year is None:
        archive_year = datetime.now().year

    if terminated_from and terminated_to and terminated_from > terminated_to:
        raise HTTPException(status_code=400, detail="terminated_from must be before or equal to terminated_to")

    filters = _build_terminated_contract_filters(
        archive_year=archive_year,
        group_id=group_id,
        search=search,
        terminated_from=terminated_from,
        terminated_to=terminated_to,
    )

    query = (
        select(Contract)
        .join(Student, Contract.student_id == Student.id)
        .options(
            selectinload(Contract.student).selectinload(Student.group),
            selectinload(Contract.group),
            selectinload(Contract.terminated_by),
            selectinload(Contract.transactions),
        )
        .where(and_(*filters))
        .order_by(Contract.terminated_at.desc().nullslast(), Contract.created_at.desc())
    )

    offset = (page - 1) * page_size
    result = await db.execute(query.offset(offset).limit(page_size))
    contracts = result.scalars().all()

    count_result = await db.execute(
        select(func.count(Contract.id))
        .join(Student, Contract.student_id == Student.id)
        .where(and_(*filters))
    )
    total = count_result.scalar() or 0

    rows = [_build_terminated_unpaid_row(contract) for contract in contracts]

    return DataResponse(
        data=rows,
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=(total + page_size - 1) // page_size,
        ),
    )


@router.get(
    "/terminated-unpaid-report/export",
    dependencies=[Depends(require_permission(PERM_CONTRACTS_VIEW))],
)
async def export_terminated_unpaid_report(
    db: Annotated[AsyncSession, Depends(get_db)],
    archive_year: int | None = Query(None, description="Filter by archive year (defaults to current year)"),
    group_id: Optional[int] = Query(None, description="Filter by group"),
    search: Optional[str] = Query(None, description="Search by contract number or student name/phone"),
    terminated_from: Optional[date] = Query(None, description="Termination start date (YYYY-MM-DD)"),
    terminated_to: Optional[date] = Query(None, description="Termination end date (YYYY-MM-DD)"),
):
    """
    Excel export for terminated contracts unpaid-months report.
    """
    if archive_year is None:
        archive_year = datetime.now().year

    if terminated_from and terminated_to and terminated_from > terminated_to:
        raise HTTPException(status_code=400, detail="terminated_from must be before or equal to terminated_to")

    filters = _build_terminated_contract_filters(
        archive_year=archive_year,
        group_id=group_id,
        search=search,
        terminated_from=terminated_from,
        terminated_to=terminated_to,
    )

    query = (
        select(Contract)
        .join(Student, Contract.student_id == Student.id)
        .options(
            selectinload(Contract.student).selectinload(Student.group),
            selectinload(Contract.group),
            selectinload(Contract.terminated_by),
            selectinload(Contract.transactions),
        )
        .where(and_(*filters))
        .order_by(Contract.terminated_at.desc().nullslast(), Contract.created_at.desc())
    )
    result = await db.execute(query)
    contracts = result.scalars().all()
    rows = [_build_terminated_unpaid_row(contract) for contract in contracts]

    wb = Workbook()
    ws = wb.active
    ws.title = "Terminated Unpaid Report"

    headers = [
        "Student ID",
        "First Name",
        "Last Name",
        "Phone",
        "Contract Number",
        "Original Contract Number",
        "Contract Group",
        "Contract Group Identifier",
        "Current Student Group",
        "Contract Start",
        "Contract End",
        "Terminated At",
        "Terminated By",
        "Termination Reason",
        "Paid Months",
        "Unpaid Months",
        "Expected Months",
        "Paid Months Count",
        "Unpaid Months Count",
        "Total Expected",
        "Total Paid",
        "Debt Amount",
    ]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_num, row in enumerate(rows, 2):
        ws.cell(row=row_num, column=1, value=row.student_id)
        ws.cell(row=row_num, column=2, value=row.student_first_name)
        ws.cell(row=row_num, column=3, value=row.student_last_name)
        ws.cell(row=row_num, column=4, value=row.student_phone or "")
        ws.cell(row=row_num, column=5, value=row.contract_number)
        ws.cell(row=row_num, column=6, value=row.original_contract_number or "")
        ws.cell(row=row_num, column=7, value=row.contract_group_name or "")
        ws.cell(row=row_num, column=8, value=row.contract_group_identifier or "")
        ws.cell(row=row_num, column=9, value=row.current_student_group_name or "")
        ws.cell(row=row_num, column=10, value=row.contract_start_date.isoformat())
        ws.cell(row=row_num, column=11, value=row.contract_end_date.isoformat())
        ws.cell(row=row_num, column=12, value=row.terminated_at.isoformat() if row.terminated_at else "")
        ws.cell(row=row_num, column=13, value=row.terminated_by_full_name or "")
        ws.cell(row=row_num, column=14, value=row.termination_reason or "")
        ws.cell(row=row_num, column=15, value=", ".join(row.paid_months) if row.paid_months else "")
        ws.cell(row=row_num, column=16, value=", ".join(row.unpaid_months) if row.unpaid_months else "")
        ws.cell(row=row_num, column=17, value=row.expected_months_count)
        ws.cell(row=row_num, column=18, value=row.paid_months_count)
        ws.cell(row=row_num, column=19, value=row.unpaid_months_count)
        ws.cell(row=row_num, column=20, value=row.total_expected)
        ws.cell(row=row_num, column=21, value=row.total_paid)
        ws.cell(row=row_num, column=22, value=row.debt_amount)

    for column_cells in ws.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    summary_row = len(rows) + 3
    ws.cell(row=summary_row, column=1, value="TOTALS").font = Font(bold=True)
    ws.cell(row=summary_row, column=21, value=sum(r.total_expected for r in rows)).font = Font(bold=True)
    ws.cell(row=summary_row, column=22, value=sum(r.total_paid for r in rows)).font = Font(bold=True)
    ws.cell(row=summary_row, column=23, value=sum(r.debt_amount for r in rows)).font = Font(bold=True)

    ws.cell(row=summary_row + 2, column=1, value=f"Total Contracts: {len(rows)}").font = Font(italic=True)
    ws.cell(
        row=summary_row + 3,
        column=1,
        value=f"Contracts With Debt: {sum(1 for r in rows if r.debt_amount > 0.01)}"
    ).font = Font(italic=True)

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"terminated_unpaid_report_{archive_year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ascii_filename = re.sub(r"[^A-Za-z0-9._-]", "_", filename)
    encoded_filename = quote(filename)
    content_disposition = (
        f"attachment; filename=\"{ascii_filename}\"; "
        f"filename*=UTF-8''{encoded_filename}"
    )

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": content_disposition},
    )


'''

@router.get("/{contract_id}", response_model=DataResponse[ContractRead], dependencies=[Depends(require_permission(PERM_CONTRACTS_VIEW))])
async def get_contract(
    contract_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(
        select(Contract)
        .options(selectinload(Contract.terminated_by))
        .where(Contract.id == contract_id)
    )
    contract = result.scalar_one_or_none()

    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")

    return DataResponse(data=ContractRead.model_validate(contract))


@router.patch("/{contract_id}", response_model=DataResponse[ContractRead], dependencies=[Depends(require_permission(PERM_CONTRACTS_EDIT))])
async def update_contract(
    contract_id: int,
    data: ContractUpdate,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from app.models.enums import ContractStatus

    result = await db.execute(select(Contract).where(Contract.id == contract_id))
    contract = result.scalar_one_or_none()

    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")

    update_data = data.model_dump(exclude_unset=True)

    # Check for duplicate contract number if it's being updated
    if "contract_number" in update_data:
        existing_contract = await db.execute(
            select(Contract).where(
                Contract.contract_number == update_data["contract_number"],
                Contract.id != contract_id
            )
        )
        if existing_contract.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="This contract already exists")

    # Check if student already has an active contract (when changing status to ACTIVE or changing student)
    if "status" in update_data and update_data["status"] == ContractStatus.ACTIVE:
        active_contract = await db.execute(
            select(Contract).where(
                Contract.student_id == contract.student_id,
                Contract.status == ContractStatus.ACTIVE,
                Contract.id != contract_id
            )
        )
        if active_contract.scalar_one_or_none():
            raise HTTPException(
                status_code=400,
                detail="This student already has an active contract. A student can only have one active contract at a time"
            )

    if "student_id" in update_data and update_data["student_id"] != contract.student_id:
        # Check if the new student already has an active contract (only if this contract is active)
        if contract.status == ContractStatus.ACTIVE:
            active_contract = await db.execute(
                select(Contract).where(
                    Contract.student_id == update_data["student_id"],
                    Contract.status == ContractStatus.ACTIVE
                )
            )
            if active_contract.scalar_one_or_none():
                raise HTTPException(
                    status_code=400,
                    detail="The target student already has an active contract. A student can only have one active contract at a time"
                )

    for field, value in update_data.items():
        setattr(contract, field, value)

    await db.commit()
    await db.refresh(contract)
    return DataResponse(data=ContractRead.model_validate(contract))


@router.patch("/{contract_id}/dates", response_model=DataResponse[ContractRead], dependencies=[Depends(require_permission(PERM_CONTRACTS_EDIT))])
async def update_contract_dates(
    contract_id: int,
    data: ContractDatesUpdate,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Update only contract start_date and/or end_date.
    """
    result = await db.execute(select(Contract).where(Contract.id == contract_id))
    contract = result.scalar_one_or_none()

    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")

    if data.start_date is None and data.end_date is None:
        raise HTTPException(
            status_code=400,
            detail="At least one field is required: start_date or end_date",
        )

    new_start_date = data.start_date if data.start_date is not None else contract.start_date
    new_end_date = data.end_date if data.end_date is not None else contract.end_date

    if new_start_date > new_end_date:
        raise HTTPException(
            status_code=400,
            detail="start_date cannot be later than end_date",
        )

    contract.start_date = new_start_date
    contract.end_date = new_end_date

    await db.commit()
    await db.refresh(contract)
    return DataResponse(data=ContractRead.model_validate(contract))


@router.post("/{contract_id}/terminate", response_model=DataResponse[ContractRead])
async def terminate_contract(
    contract_id: int,
    data: ContractTerminate,
    user: Annotated[User, Depends(require_permission(PERM_CONTRACTS_EDIT))],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Terminate a contract with reason and termination date.
    Automatically changes contract status to TERMINATED and renames the contract number to:
    {old_number}-T-yyyymmdd
    Records who terminated the contract with their full name.
    """
    result = await db.execute(select(Contract).where(Contract.id == contract_id))
    contract = result.scalar_one_or_none()

    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")

    if contract.terminated_at or contract.status == ContractStatus.TERMINATED:
        raise HTTPException(status_code=400, detail="Contract is already terminated")

    termination_dt = data.terminated_at or datetime.utcnow()
    terminated_contract_number = await _generate_unique_terminated_contract_number(
        db=db,
        base_contract_number=contract.contract_number,
        termination_dt=termination_dt,
        current_contract_id=contract.id,
    )

    # Set termination details
    contract.terminated_at = termination_dt
    contract.terminated_by_user_id = user.id
    contract.termination_reason = data.termination_reason
    contract.status = ContractStatus.TERMINATED
    contract.contract_number = terminated_contract_number

    # Student should not remain active after contract termination.
    student_result = await db.execute(select(Student).where(Student.id == contract.student_id))
    student = student_result.scalar_one_or_none()
    if student:
        student.status = StudentStatus.INACTIVE

    await db.commit()

    # Refresh with terminated_by relationship
    await db.refresh(contract, ["terminated_by"])

    return DataResponse(data=ContractRead.model_validate(contract))


@router.patch("/{contract_id}/update-pdf", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_CONTRACTS_EDIT))])
async def update_contract_pdf(
    contract_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    final_pdf: UploadFile = File(..., description="New PDF file to replace the existing contract PDF"),
):
    """
    Update only the final PDF of an existing contract.

    This endpoint allows you to replace the contract's PDF without modifying any other data.
    The old PDF URL in the database will be replaced with the new one.

    **Note**: The old PDF file in S3 will remain (not deleted) but the database will point to the new PDF.

    Args:
        contract_id: The ID of the contract to update
        final_pdf: New PDF file to upload

    Returns:
        Success message with new PDF URL
    """
    from app.core.s3 import upload_pdf_to_s3
    import tempfile

    # Find contract
    result = await db.execute(select(Contract).where(Contract.id == contract_id))
    contract = result.scalar_one_or_none()

    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")

    # Save old PDF URL for response
    old_pdf_url = contract.final_pdf_url

    # Save uploaded PDF to temporary file
    try:
        # Create temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
            content = await final_pdf.read()
            temp_pdf.write(content)
            temp_pdf_path = temp_pdf.name

        # Upload new PDF to S3
        new_pdf_url = await upload_pdf_to_s3(temp_pdf_path, contract.contract_number)

        # Clean up temp file
        try:
            os.unlink(temp_pdf_path)
        except:
            pass

    except Exception as e:
        # Clean up temp file on error
        try:
            if 'temp_pdf_path' in locals():
                os.unlink(temp_pdf_path)
        except:
            pass
        raise HTTPException(status_code=500, detail=f"Failed to upload PDF to S3: {str(e)}")

    # Update contract with new PDF URL
    contract.final_pdf_url = new_pdf_url
    await db.commit()
    await db.refresh(contract)

    return DataResponse(data={
        "message": "Contract PDF updated successfully",
        "contract_id": contract_id,
        "contract_number": contract.contract_number,
        "old_pdf_url": old_pdf_url,
        "new_pdf_url": new_pdf_url
    })


'''
@router.get("/payment-months/{contract_number}", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_CONTRACTS_VIEW))])
async def get_contract_payment_months(
    contract_number: str,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Get valid payment months for a contract based on start date and end/termination date.

    Uses contract_number instead of contract_id.

    Payment months are calculated:
    - Starting from the month of contract start_date
    - Ending at the month of contract end_date or terminated_at (whichever is earlier)
    - Only includes months within the contract period
    """
    result = await db.execute(select(Contract).where(Contract.contract_number == contract_number))
    contract = result.scalar_one_or_none()

    if not contract:
        raise HTTPException(status_code=404, detail=f"Contract '{contract_number}' not found")

    # Determine the effective end date (earliest of end_date or terminated_at)
    effective_end_date = contract.end_date
    if contract.terminated_at:
        termination_date = contract.terminated_at.date()
        if termination_date < effective_end_date:
            effective_end_date = termination_date

    # Generate list of valid payment months
    payment_months = []
    current_date = contract.start_date.replace(day=1)  # Start from first day of start month

    while current_date <= effective_end_date:
        payment_months.append({
            "year": current_date.year,
            "month": current_date.month,
            "month_name": current_date.strftime("%B"),
            "display": current_date.strftime("%B %Y")
        })

        # Move to next month
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)

    return DataResponse(data={
        "contract_id": contract.id,
        "contract_number": contract.contract_number,
        "start_date": str(contract.start_date),
        "end_date": str(contract.end_date),
        "terminated_at": str(contract.terminated_at.date()) if contract.terminated_at else None,
        "effective_end_date": str(effective_end_date),
        "payment_months": payment_months,
        "total_months": len(payment_months)
    })


'''

@router.delete("/{contract_id}", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_CONTRACTS_EDIT))])
async def delete_contract(
    contract_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Soft delete a contract by setting its status to DELETED.
    The contract is not actually removed from the database.
    """
    result = await db.execute(select(Contract).where(Contract.id == contract_id))
    contract = result.scalar_one_or_none()

    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")

    # Soft delete: set status to DELETED instead of actually deleting
    contract.status = ContractStatus.DELETED
    await db.commit()

    return DataResponse(data={"message": "Contract deleted successfully"})


@router.post("/bulk-delete", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_CONTRACTS_EDIT))])
async def bulk_delete_contracts(
    contract_ids: list[int],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Soft delete multiple contracts by setting their status to DELETED.
    Contracts are not actually removed from the database.
    """
    if not contract_ids:
        raise HTTPException(status_code=400, detail="No contract IDs provided")

    deleted_count = 0
    errors = []

    for contract_id in contract_ids:
        try:
            result = await db.execute(select(Contract).where(Contract.id == contract_id))
            contract = result.scalar_one_or_none()

            if not contract:
                errors.append({"contract_id": contract_id, "error": "Contract not found"})
                continue

            # Soft delete: set status to DELETED instead of actually deleting
            contract.status = ContractStatus.DELETED
            deleted_count += 1
        except Exception as e:
            errors.append({"contract_id": contract_id, "error": str(e)})

    await db.commit()

    return DataResponse(data={
        "message": f"Deleted {deleted_count} contract(s)",
        "deleted_count": deleted_count,
        "total_requested": len(contract_ids),
        "errors": errors if errors else None
    })



@router.post("/create-with-files", response_class=FileResponse, deprecated=True)
async def create_contract_with_file_upload(
    user: Annotated[User, Depends(require_permission(PERM_CONTRACTS_EDIT))],
    db: Annotated[AsyncSession, Depends(get_db)],
    student_id: int = Form(...),
    group_id: int = Form(...),
    contract_number: str = Form(...),
    custom_fields_json: str = Form(...),
    passport_copy: UploadFile = File(...),
    form_086: UploadFile = File(...),
    heart_checkup: UploadFile = File(...),
    birth_certificate: UploadFile = File(...),
    contract_images: List[UploadFile] = File(...)
):
    """
    ⚠️ DEPRECATED: This endpoint is deprecated and disabled.

    Please use POST /students/create-with-contract instead.

    The new endpoint:
    - Creates both student AND contract in one operation
    - Has all form fields visible (not hidden in JSON)
    - Uploads files to S3 automatically
    - Generates and returns PDF contract
    """
    raise HTTPException(
        status_code=410,
        detail="This endpoint is deprecated. Please use POST /students/create-with-contract instead."
    )
    # Parse custom fields
    try:
        custom_fields_dict = json.loads(custom_fields_json)
        custom_fields = ContractCustomFields(**custom_fields_dict)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid custom_fields JSON: {e}")

    # Validate student
    student_result = await db.execute(select(Student).where(Student.id == student_id))
    student = student_result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail=f"Student ID {student_id} not found")

    # Check for active contract
    active_contract = await db.execute(
        select(Contract).where(
            Contract.student_id == student_id,
            Contract.status == ContractStatus.ACTIVE
        )
    )
    if active_contract.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Student already has active contract")

    # Group check
    group_result = await db.execute(select(Group).where(Group.id == group_id))
    group = group_result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail=f"Group ID {group_id} not found")

    birth_year = custom_fields.student.birth_year

    # Check for duplicate contract number
    existing_contract = await db.execute(
        select(Contract).where(Contract.contract_number == contract_number)
    )
    if existing_contract.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Contract number already exists")

    # Parse contract number: N{identifier}{seq}{year}
    import re
    if not contract_number.startswith("N"):
        raise HTTPException(status_code=400, detail="Contract number must start with 'N'")

    # Extract year (last 4 digits)
    if not contract_number.endswith(str(birth_year)):
        raise HTTPException(status_code=400, detail=f"Contract number must end with birth year {birth_year}")

    # Extract middle part (identifier + sequence)
    middle = contract_number[1:-len(str(birth_year))]

    # Try to match pattern: identifier (letters/numbers) + sequence (numbers only)
    match = re.match(r'^(.+?)(\d+)$', middle)
    if not match:
        raise HTTPException(status_code=400, detail="Invalid contract number format (must be N{identifier}{seq}{year})")

    extracted_identifier = match.group(1)
    sequence_number = int(match.group(2))

    # Verify identifier matches group
    if extracted_identifier != group.identifier:
        raise HTTPException(
            status_code=400,
            detail=f"Contract number identifier '{extracted_identifier}' doesn't match group identifier '{group.identifier}'"
        )

    # Upload images to S3
    try:
        passport_url = upload_image_to_s3(passport_copy, folder="contracts/passports")
        form_086_url = upload_image_to_s3(form_086, folder="contracts/medical")
        heart_url = upload_image_to_s3(heart_checkup, folder="contracts/medical")
        birth_cert_url = upload_image_to_s3(birth_certificate, folder="contracts/birth_certificates")
        contract_image_urls = [upload_image_to_s3(img, folder="contracts/pages") for img in contract_images]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to upload files to S3: {e}")

    # Create contract
    contract = Contract(
        contract_number=contract_number,
        birth_year=birth_year,
        sequence_number=sequence_number,
        start_date=custom_fields.contract_terms.contract_start_date,
        end_date=custom_fields.contract_terms.contract_end_date,
        status=ContractStatus.ACTIVE,
        student_id=student_id,
        group_id=group_id,
        passport_copy_url=passport_url,
        form_086_url=form_086_url,
        heart_checkup_url=heart_url,
        birth_certificate_url=birth_cert_url,
        contract_images_urls=json.dumps(contract_image_urls),
        custom_fields=json.dumps(custom_fields.model_dump(), ensure_ascii=False, default=str)
    )

    db.add(contract)
    await db.commit()
    await db.refresh(contract)

    # Generate PDF
    try:
        # Create temporary directory if not exists
        os.makedirs("temp_pdfs", exist_ok=True)

        pdf_path = f"temp_pdfs/contract_{contract.contract_number}_{contract.student_id}_{contract.birth_year}.pdf"

        # Prepare PDF data
        pdf_data = {
            "shartnoma_raqami": contract.contract_number,
            "sana": {
                "kun": f"{contract.start_date.day:02d}",
                "oy": contract.start_date.strftime('%B'),
                "yil": str(contract.start_date.year)
            },
            "buyurtmachi": {
                "fio": custom_fields.customer.full_name,
                "pasport_seriya": custom_fields.customer.passport_number,
                "pasport_kim_bergan": custom_fields.customer.passport_issued_by,
                "pasport_qachon_bergan": str(custom_fields.customer.passport_issue_date),
                "manzil": custom_fields.customer.address,
                "telefon": custom_fields.customer.phone
            },
            "tarbiyalanuvchi": {
                "fio": f"{custom_fields.student.first_name} {custom_fields.student.last_name}",
                "tugilganlik_guvohnoma": custom_fields.student_birth_certificate.series,
                "guvohnoma_kim_bergan": custom_fields.student_birth_certificate.issued_by,
                "guvohnoma_qachon_bergan": str(custom_fields.student_birth_certificate.issue_date)
            },
            "shartnoma_muddati": {
                "boshlanish": contract.start_date.strftime('«%d» %B'),
                "tugash": contract.end_date.strftime('«%d» %B'),
                "yil": str(contract.start_date.year)
            }
        }

        pdf_data_cleaned = convert_dates(pdf_data)

        # Generate PDF
        generator = ContractGenerator(pdf_data_cleaned)
        pdf_generated_path = generator.generate(pdf_path)

        return FileResponse(
            path=pdf_generated_path,
            media_type="application/pdf",
            filename=os.path.basename(pdf_generated_path)
        )

    except Exception as e:
        # Rollback contract creation if PDF generation fails
        await db.delete(contract)
        await db.commit()
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {e}")


@router.get("/next-available/{group_id}", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_CONTRACTS_VIEW))])
async def get_next_available_number(
    group_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Get the next available contract number for a group.

    Birth year is automatically taken from the group's birth_year.
    Returns the next sequential contract number to use.

    Example: If group B1 (birth year 2020) has 1-2020B1, 2-2020B1, 3-2020B1 used, returns 4-2020B1.
    """
    # Get group
    group_result = await db.execute(select(Group).where(Group.id == group_id))
    group = group_result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail=f"Group with ID {group_id} not found")

    # Use group's birth_year
    birth_year = group.birth_year

    # Get available numbers
    try:
        available_numbers = await get_available_contract_numbers(db, group_id, birth_year)
    except ContractNumberAllocationError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if not available_numbers:
        return DataResponse(data={
            "next_number": None,
            "contract_number": None,
            "message": f"Group '{group.name}' is full (capacity: {group.capacity})",
            "is_full": True,
            "group_name": group.name,
            "birth_year": birth_year
        })

    # Get the first (lowest) available number
    next_seq = available_numbers[0]
    contract_number = f"{next_seq}-{birth_year}{group.identifier}"

    return DataResponse(data={
        "next_number": next_seq,
        "contract_number": contract_number,
        "message": f"Next available number for group '{group.name}' ({group.identifier})",
        "is_full": False,
        "group_name": group.name,
        "group_identifier": group.identifier,
        "birth_year": birth_year,
        "group_capacity": group.capacity,
        "total_used": group.capacity - len(available_numbers)
    })


@router.get("/available-numbers/{group_id}", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_CONTRACTS_VIEW))])
async def get_all_available_numbers(
    group_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Get available contract numbers (ONLY GAPS - not unused numbers after max).

    **Returns only GAP numbers** - numbers that are missing in the sequence.

    Excludes numbers used by:
    - ACTIVE contracts
    - EXPIRED contracts
    - ARCHIVED contracts
    - DELETED contracts

    Note:
    - TERMINATED contracts are considered released and their numbers can be reused.

    **Only returns gaps in the sequence:**
    - If contracts exist: 1,2,4,5 → Returns: [3] (only the gap)
    - Does NOT return numbers after max (6,7,8... are not returned)
    - Empty result if sequence is continuous (no gaps)

    Example 1:
    - Existing: 1-2020B1, 2-2020B1, 4-2020B1, 5-2020B1
    - Contract 3-2020B1 was deleted
    - Returns: [3] (only the gap between 2 and 4)

    Example 2:
    - Existing: 1-2020B1, 2-2020B1, 3-2020B1
    - No gaps in sequence
    - Returns: [] (empty - no gaps)

    Example 3:
    - Existing: 1-2020B1, 3-2020B1, 6-2020B1
    - Gaps: 2 (between 1-3), 4,5 (between 3-6)
    - Returns: [2, 4, 5]
    """
    from app.models.enums import ContractStatus

    # Get group
    group_result = await db.execute(select(Group).where(Group.id == group_id))
    group = group_result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail=f"Group with ID {group_id} not found")

    birth_year = group.birth_year

    # Get all USED sequence numbers that remain reserved.
    # TERMINATED contracts are intentionally excluded (numbers are reusable).
    # These numbers are NOT available for reuse
    used_contracts = await db.execute(
        select(Contract.sequence_number).where(
            and_(
                Contract.group_id == group_id,
                Contract.birth_year == birth_year,
                Contract.status.in_([
                    ContractStatus.ACTIVE,
                    ContractStatus.EXPIRED,
                    ContractStatus.ARCHIVED,
                    ContractStatus.DELETED,
                ])
            )
        ).order_by(Contract.sequence_number)
    )
    used_numbers_list = sorted([row[0] for row in used_contracts.fetchall()])

    # Find gaps ONLY (not numbers after max)
    gap_numbers = []
    if used_numbers_list:
        # Check for gaps between min and max used numbers
        min_used = used_numbers_list[0]
        max_used = used_numbers_list[-1]

        for num in range(min_used, max_used + 1):
            if num not in used_numbers_list:
                gap_numbers.append(num)

    # Build contract number strings for gaps only
    gap_contract_numbers = [
        f"{seq}-{birth_year}{group.identifier}" for seq in gap_numbers
    ]

    return DataResponse(data={
        "group_id": group_id,
        "group_name": group.name,
        "group_identifier": group.identifier,
        "birth_year": birth_year,
        "capacity": group.capacity,
        "used_count": len(used_numbers_list),
        "gap_count": len(gap_numbers),
        "gap_sequence_numbers": gap_numbers,
        "gap_contract_numbers": gap_contract_numbers,
        "used_sequence_numbers": used_numbers_list,
        "max_used_number": used_numbers_list[-1] if used_numbers_list else 0,
        "has_gaps": len(gap_numbers) > 0
    })


import io
import httpx
from fastapi.responses import StreamingResponse
from fastapi.responses import JSONResponse

@router.get("/{year}/{contract_number}/pdf", dependencies=[Depends(require_permission(PERM_CONTRACTS_VIEW))])
async def get_contract_pdf_url(
    year: int,
    contract_number: str,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Get contract PDF URL by year and contract number.
    Instead of returning the actual PDF stream, this endpoint returns the S3 URL.

    Example:
        GET /contracts/2025/1-2017B1/pdf
        → Returns: {"pdf_url": "https://bucket.s3.region.amazonaws.com/contract-pdfs/1-2017B1_xxx.pdf"}
    """

    # Find contract in DB
    result = await db.execute(
        select(Contract).where(
            and_(
                Contract.contract_number == contract_number,
                Contract.archive_year == year
            )
        )
    )
    contract = result.scalar_one_or_none()

    if not contract:
        raise HTTPException(
            status_code=404,
            detail=f"Contract {contract_number} for year {year} not found"
        )

    if not contract.final_pdf_url:
        raise HTTPException(
            status_code=404,
            detail=f"PDF not generated for contract {contract_number} yet"
        )

    # Return JSON with the S3 URL
    return JSONResponse(content={"pdf_url": contract.final_pdf_url})


'''
@router.get("/{contract_number}/payment-status", response_model=DataResponse[ContractPaymentStatus], dependencies=[Depends(require_permission(PERM_CONTRACTS_VIEW))])
async def get_contract_payment_status(
    contract_number: str,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Get payment status for a contract - shows paid and unpaid months.

    Returns:
    - Contract details
    - List of paid months with transaction details
    - List of unpaid months with expected amounts
    - Total paid, unpaid, and expected amounts

    Example:
        GET /contracts/1-2020B1/payment-status
    """
    from datetime import date as date_class
    from sqlalchemy.dialects.postgresql import JSONB
    from sqlalchemy import cast

    # Find contract
    result = await db.execute(
        select(Contract).options(selectinload(Contract.student)).where(
            Contract.contract_number == contract_number
        )
    )
    contract = result.scalar_one_or_none()

    if not contract:
        raise HTTPException(
            status_code=404,
            detail=f"Contract {contract_number} not found"
        )

    # Get student
    if not contract.student:
        raise HTTPException(
            status_code=404,
            detail=f"Student not found for contract {contract_number}"
        )

    student = contract.student

    # Calculate all months in effective contract period.
    # If terminated, months after terminated_at are excluded.
    effective_end_date = contract.end_date
    if contract.terminated_at:
        termination_date = contract.terminated_at.date()
        if termination_date < effective_end_date:
            effective_end_date = termination_date

    contract_start_month = date_class(contract.start_date.year, contract.start_date.month, 1)
    contract_end_month = date_class(effective_end_date.year, effective_end_date.month, 1)

    # Generate list of all months in contract period
    all_months = []
    current_month = contract_start_month
    while current_month <= contract_end_month:
        all_months.append((current_month.year, current_month.month))
        # Move to next month
        if current_month.month == 12:
            current_month = date_class(current_month.year + 1, 1, 1)
        else:
            current_month = date_class(current_month.year, current_month.month + 1, 1)

    # Get all successful transactions for this contract
    transactions_result = await db.execute(
        select(Transaction).where(
            Transaction.contract_id == contract.id,
            Transaction.status == PaymentStatus.SUCCESS
        ).order_by(Transaction.payment_year, Transaction.paid_at)
    )
    transactions = transactions_result.scalars().all()

    # Build paid month details with per-month allocation and settlement status.
    all_months_set = set(all_months)
    paid_months_map: dict[tuple[int, int], dict] = {}
    for transaction in transactions:
        if transaction.payment_year is None:
            continue

        months = normalize_unique_months(transaction.payment_months)
        if not months:
            continue

        per_month_amount = Decimal(str(transaction.amount)) / Decimal(len(months))
        for month in months:
            key = (int(transaction.payment_year), month)
            if key not in all_months_set:
                continue

            if key not in paid_months_map:
                paid_months_map[key] = {
                    "year": int(transaction.payment_year),
                    "month": month,
                    "amount": Decimal("0"),
                    "paid_at": transaction.paid_at,
                    "source": transaction.source,
                    "transaction_id": transaction.id,
                    "has_waiver": False,
                    "has_payment": False,
                    "comments": [],
                }

            month_entry = paid_months_map[key]
            month_entry["amount"] += per_month_amount

            if transaction.paid_at and (
                month_entry["paid_at"] is None or transaction.paid_at > month_entry["paid_at"]
            ):
                month_entry["paid_at"] = transaction.paid_at
                month_entry["source"] = transaction.source
                month_entry["transaction_id"] = transaction.id

            if transaction.settlement_type == PaymentSettlementType.WAIVER_SPRAVKA:
                month_entry["has_waiver"] = True
            else:
                month_entry["has_payment"] = True

            comment = (transaction.comment or "").strip()
            if comment and comment not in month_entry["comments"]:
                month_entry["comments"].append(comment)

    # Separate paid and unpaid months
    paid_months = []
    unpaid_months = []

    for year, month in all_months:
        if (year, month) in paid_months_map:
            month_entry = paid_months_map[(year, month)]
            if month_entry["has_waiver"] and not month_entry["has_payment"]:
                settlement_type = PaymentSettlementType.WAIVER_SPRAVKA
                settlement_status = "waiver_spravka"
            elif month_entry["has_waiver"] and month_entry["has_payment"]:
                settlement_type = PaymentSettlementType.PAYMENT
                settlement_status = "mixed"
            else:
                settlement_type = PaymentSettlementType.PAYMENT
                settlement_status = "payment"

            paid_months.append(
                PaidMonth(
                    year=year,
                    month=month,
                    amount=float(month_entry["amount"].quantize(Decimal("0.01"))),
                    status=PaymentStatus.SUCCESS,
                    paid_at=month_entry["paid_at"],
                    source=month_entry["source"],
                    transaction_id=month_entry["transaction_id"],
                    settlement_type=settlement_type,
                    settlement_status=settlement_status,
                    comment="; ".join(month_entry["comments"]) if month_entry["comments"] else None,
                )
            )
        else:
            unpaid_months.append(UnpaidMonth(
                year=year,
                month=month,
                expected_amount=0
            ))

    # Calculate totals
    total_paid = sum(pm.amount for pm in paid_months)
    total_unpaid = sum(um.expected_amount for um in unpaid_months)
    waived_only_month_count = sum(
        1
        for pm in paid_months
        if pm.settlement_status == "waiver_spravka"
    )
    total_expected = 0

    return DataResponse(data=ContractPaymentStatus(
        contract_number=contract.contract_number,
        student_name=f"{student.first_name} {student.last_name}",
        start_date=contract.start_date.isoformat(),
        end_date=effective_end_date.isoformat(),
        contract_status=contract.status.value,
        paid_months=paid_months,
        unpaid_months=unpaid_months,
        total_paid=total_paid,
        total_unpaid=total_unpaid,
        total_expected=total_expected,
    ))


'''

@router.get("/{year}/{contract_number}/student-data", dependencies=[Depends(require_permission(PERM_CONTRACTS_VIEW))])
async def get_contract_student_data(
    year: int,
    contract_number: str,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Get student data from an old contract for creating a new year contract.

    This endpoint is useful when creating a new contract for a student who had
    a contract in a previous year. It returns all student information from the
    old contract, including photos, so it can be used to pre-fill the new contract form.

    Returns:
    - Student personal information
    - Document photos (passport, form 086, etc.)
    - Old contract details

    Example:
        GET /contracts/2025/1-2020B1/student-data
    """
    # Find contract with related student data
    result = await db.execute(
        select(Contract).options(selectinload(Contract.student)).where(
            and_(
                Contract.contract_number == contract_number,
                Contract.archive_year == year
            )
        )
    )
    contract = result.scalar_one_or_none()

    if not contract:
        raise HTTPException(
            status_code=404,
            detail=f"Contract {contract_number} for year {year} not found"
        )

    if not contract.student:
        raise HTTPException(
            status_code=404,
            detail=f"Student not found for contract {contract_number}"
        )

    student = contract.student

    # Parse photo URLs from JSON strings
    def parse_photo_urls(url_json: str | None) -> list[str]:
        if not url_json:
            return []
        try:
            urls = json.loads(url_json)
            return urls if isinstance(urls, list) else [urls]
        except (json.JSONDecodeError, TypeError):
            return [url_json] if url_json else []

    # Build response with all student data
    student_data = {
        "student": {
            "id": student.id,
            "first_name": student.first_name,
            "last_name": student.last_name,
            "date_of_birth": student.date_of_birth.isoformat(),
            "height": student.height,
            "weight": student.weight,
            "ampula": student.ampula,
            "pnfl": student.pnfl,
            "phone": student.phone,
            "address": student.address,
            "photo_url": student.photo_url,
            "status": student.status.value if student.status else None,
        },
        "old_contract": {
            "contract_number": contract.contract_number,
            "start_date": contract.start_date.isoformat(),
            "end_date": contract.end_date.isoformat(),
            "status": contract.status.value,
            "archive_year": contract.archive_year,
            "birth_year": contract.birth_year,
        },
        "documents": {
            "passport_copies": parse_photo_urls(contract.passport_copy_url),
            "form_086": parse_photo_urls(contract.form_086_url),
            "other_documents": parse_photo_urls(contract.other_documents_url) if hasattr(contract, 'other_documents_url') else []
        }
    }

    return JSONResponse(content=student_data)



from fastapi.responses import RedirectResponse

from app.core.s3 import generate_signed_url



@router.get("/{contract_id}/pdf", dependencies=[Depends(require_permission(PERM_CONTRACTS_VIEW))])
async def view_contract_pdf(
    contract_id: int,
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Contract).where(Contract.id == contract_id))
    contract = result.scalar_one_or_none()

    if not contract or not contract.final_pdf_url:
        raise HTTPException(404, "PDF not found")

    signed_url = contract.final_pdf_url
    return RedirectResponse(signed_url)
