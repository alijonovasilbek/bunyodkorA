from typing import Annotated
from datetime import datetime, date
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from openpyxl import load_workbook
from io import BytesIO
from app.core.db import get_db
from app.core.permissions import PERM_STUDENTS_EDIT
from app.schemas.common import DataResponse
from app.deps import require_permission
from app.models.domain import Student, Group
from app.models.enums import StudentStatus

router = APIRouter(prefix="/import", tags=["Import"])


@router.post("/students", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def import_students(
    file: UploadFile = File(...),
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    """
    Import students from Excel file.

    Expected columns:
    - first_name: Student first name (required)
    - last_name: Student last name (required)
    - date_of_birth: Format YYYY-MM-DD (required)
    - height: Student height (required)
    - weight: Student weight (required)
    - ampula: Student ampula/position
    - millati: Student nationality
    - pnfl: Unique 14-digit PNFL (required)
    - phone: Student phone number
    - address: Student address
    - status: ACTIVE, INACTIVE, GRADUATED, EXPELLED (default: ACTIVE)
    - group_name: Name of the group to assign student to
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    try:
        # Read Excel file
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents))
        sheet = workbook.active

        # Get header row
        headers = [cell.value for cell in sheet[1]]

        success_count = 0
        error_count = 0
        errors = []

        # Process each row (skip header)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Create dictionary from row
                row_data = dict(zip(headers, row))

                # Skip empty rows
                if not row_data.get('first_name') or not row_data.get('last_name'):
                    continue

                pnfl = str(row_data.get('pnfl') or '').strip()
                if not pnfl:
                    raise ValueError(f"PNFL is required in row {row_num}")
                if len(pnfl) != 14:
                    raise ValueError(f"PNFL must be 14 characters in row {row_num}")

                height_raw = row_data.get('height')
                weight_raw = row_data.get('weight')
                if height_raw in (None, ""):
                    raise ValueError(f"Height is required in row {row_num}")
                if weight_raw in (None, ""):
                    raise ValueError(f"Weight is required in row {row_num}")

                try:
                    height = int(height_raw)
                    weight = int(weight_raw)
                except (TypeError, ValueError):
                    raise ValueError(f"Height and weight must be integers in row {row_num}")

                # Parse date_of_birth
                dob_str = row_data.get('date_of_birth')
                if isinstance(dob_str, str):
                    date_of_birth = datetime.strptime(dob_str, '%Y-%m-%d').date()
                elif isinstance(dob_str, datetime):
                    date_of_birth = dob_str.date()
                elif isinstance(dob_str, date):
                    date_of_birth = dob_str
                else:
                    raise ValueError(f"Invalid date_of_birth format in row {row_num}")

                existing_pnfl = await db.execute(
                    select(Student).where(Student.pnfl == pnfl)
                )
                if existing_pnfl.scalar_one_or_none():
                    raise ValueError(f"PNFL {pnfl} already exists")

                # Get or find group
                group_id = None
                if row_data.get('group_name'):
                    group_result = await db.execute(
                        select(Group).where(Group.name == row_data['group_name'])
                    )
                    group = group_result.scalar_one_or_none()
                    if group:
                        group_id = group.id
                    else:
                        raise ValueError(f"Group '{row_data['group_name']}' not found")

                # Create student
                student_status = row_data.get('status', 'ACTIVE')
                if isinstance(student_status, str):
                    try:
                        student_status = StudentStatus[student_status.upper()]
                    except KeyError:
                        student_status = StudentStatus.ACTIVE

                student = Student(
                    first_name=row_data['first_name'],
                    last_name=row_data['last_name'],
                    date_of_birth=date_of_birth,
                    height=height,
                    weight=weight,
                    ampula=row_data.get('ampula'),
                    millati=row_data.get('millati'),
                    pnfl=pnfl,
                    phone=row_data.get('phone'),
                    address=row_data.get('address'),
                    status=student_status,
                    archive_year=datetime.now().year,
                    group_id=group_id,
                )
                db.add(student)
                await db.flush()  # Get student ID

                await db.commit()
                success_count += 1

            except Exception as e:
                await db.rollback()
                error_count += 1
                errors.append({
                    "row": row_num,
                    "error": str(e)
                })

        return DataResponse(
            data={
                "message": "Student import completed",
                "filename": file.filename,
                "success_count": success_count,
                "error_count": error_count,
                "errors": errors if errors else None
            }
        )

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing Excel file: {str(e)}")
