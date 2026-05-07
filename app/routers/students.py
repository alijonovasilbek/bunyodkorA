from typing import Annotated, Optional, List
import re
from io import BytesIO
from urllib.parse import quote
from fastapi import APIRouter, Depends, HTTPException, Query, File, Form, UploadFile
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_, extract
from sqlalchemy.orm import selectinload
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.core.db import get_db
from app.core.permissions import PERM_STUDENTS_VIEW, PERM_STUDENTS_EDIT, PERM_ATTENDANCE_VIEW
from app.core.private_files import verify_private_file_token
from app.models.domain import Student
from app.models.attendance import Attendance, GateLog, Session
from app.models.enums import StudentStatus
from app.schemas.student import StudentRead, StudentCreate, StudentUpdate, StudentFullInfo
from app.schemas.attendance import AttendanceRead, GateLogRead
from app.schemas.common import DataResponse, PaginationMeta
from app.deps import require_permission, CurrentUser
from app.models.auth import User
from app.core.s3 import download_s3_object, upload_image_to_s3, upload_private_file_to_s3

router = APIRouter(prefix="/students", tags=["Students"])


def serialize_student(student: Student) -> StudentRead:
    return StudentRead.model_validate(student)


def raise_student_file_upload_error(exc: Exception) -> None:
    detail = str(exc)
    lowered = detail.lower()
    if "unsupported" in lowered or "validation" in lowered or "invalid" in lowered:
        raise HTTPException(status_code=400, detail=detail) from exc
    raise HTTPException(status_code=500, detail=detail) from exc


def build_student_create_form(
    first_name: str = Form(..., description="Student first name"),
    last_name: str = Form(..., description="Student last name"),
    date_of_birth: date = Form(..., description="Date of birth in YYYY-MM-DD format"),
    height: int = Form(..., description="Height in centimeters"),
    weight: int = Form(..., description="Weight in kilograms"),
    ampula: str | None = Form(None, description="Student ampula/position"),
    millati: str | None = Form(None, description="Student nationality"),
    pnfl: str = Form(..., min_length=14, max_length=14, description="14-digit PNFL"),
    phone: str | None = Form(None, description="Phone number"),
    address: str | None = Form(None, description="Address"),
    group_id: int | None = Form(None, description="Group ID"),
) -> StudentCreate:
    payload = {
        "first_name": first_name,
        "last_name": last_name,
        "date_of_birth": date_of_birth,
        "height": height,
        "weight": weight,
        "pnfl": pnfl,
        "status": StudentStatus.ACTIVE,
    }
    optional_fields = {
        "ampula": ampula,
        "millati": millati,
        "phone": phone,
        "address": address,
        "group_id": group_id,
    }
    payload.update({key: value for key, value in optional_fields.items() if value is not None})
    return StudentCreate(**payload)


def build_student_update_form(
    first_name: str | None = Form(None, description="Student first name"),
    last_name: str | None = Form(None, description="Student last name"),
    date_of_birth: date | None = Form(None, description="Date of birth in YYYY-MM-DD format"),
    height: int | None = Form(None, description="Height in centimeters"),
    weight: int | None = Form(None, description="Weight in kilograms"),
    ampula: str | None = Form(None, description="Student ampula/position"),
    millati: str | None = Form(None, description="Student nationality"),
    pnfl: str | None = Form(None, min_length=14, max_length=14, description="14-digit PNFL"),
    phone: str | None = Form(None, description="Phone number"),
    address: str | None = Form(None, description="Address"),
    status: StudentStatus | None = Form(None, description="Student status"),
    group_id: int | None = Form(None, description="Group ID"),
) -> StudentUpdate:
    payload = {
        "first_name": first_name,
        "last_name": last_name,
        "date_of_birth": date_of_birth,
        "height": height,
        "weight": weight,
        "ampula": ampula,
        "millati": millati,
        "pnfl": pnfl,
        "phone": phone,
        "address": address,
        "status": status,
        "group_id": group_id,
    }
    return StudentUpdate(**{key: value for key, value in payload.items() if value is not None})


@router.get("/search", response_model=DataResponse[list[StudentRead]], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def search_students(
    db: Annotated[AsyncSession, Depends(get_db)],
    query: str = Query(..., description="Search by first name, last name, ampula, millati, or phone"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=2000),
):
    """
    Comprehensive search for students by:
    - First name
    - Last name
    - Ampula
    - Millati
    - Phone number
    """
    students_query = select(Student).where(
        or_(
            Student.first_name.ilike(f"%{query}%"),
            Student.last_name.ilike(f"%{query}%"),
            Student.ampula.ilike(f"%{query}%"),
            Student.millati.ilike(f"%{query}%"),
            Student.phone.ilike(f"%{query}%"),
        )
    ).distinct()

    # Apply pagination
    offset = (page - 1) * page_size
    result = await db.execute(students_query.offset(offset).limit(page_size))
    students = result.scalars().all()

    # Count total results
    count_query = select(func.count(Student.id.distinct())).where(
        or_(
            Student.first_name.ilike(f"%{query}%"),
            Student.last_name.ilike(f"%{query}%"),
            Student.ampula.ilike(f"%{query}%"),
            Student.millati.ilike(f"%{query}%"),
            Student.phone.ilike(f"%{query}%"),
        )
    )
    count_result = await db.execute(count_query)
    total = count_result.scalar() or 0

    return DataResponse(
        data=[serialize_student(s) for s in students],
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=(total + page_size - 1) // page_size if total > 0 else 0,
        ),
    )


@router.get("", response_model=DataResponse[list[StudentRead]], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_students(
    db: Annotated[AsyncSession, Depends(get_db)],
    search: Optional[str] = None,
    group_id: Optional[int] = None,
    status: Optional[str] = None,
    archive_year: int | None = Query(None, description="Filter by archive year (defaults to current year)"),
    include_archived: bool = Query(False, description="Include archived students (default: only non-ARCHIVED)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=2000),
):
    """
    Get all students with optional filters.

    Default behavior:
    - Shows current year's students only
    - Shows only ACTIVE students unless filters are provided
    """
    current_year = datetime.now().year
    if archive_year is None:
        year_filter = or_(
            Student.archive_year == current_year,
            extract("year", Student.created_at) == current_year,
        )
    else:
        year_filter = Student.archive_year == archive_year

    query = select(Student).where(year_filter)

    # Status behavior:
    # - explicit status filter: return only that status
    # - include_archived=true: return all except DELETED
    # - default: ACTIVE only
    if status:
        query = query.where(Student.status == status)
    elif include_archived:
        query = query.where(Student.status != StudentStatus.DELETED)
    else:
        query = query.where(Student.status == StudentStatus.ACTIVE)

    if search:
        query = query.where(
            or_(
                Student.first_name.ilike(f"%{search}%"),
                Student.last_name.ilike(f"%{search}%"),
                Student.ampula.ilike(f"%{search}%"),
            )
        )
    if group_id:
        query = query.where(Student.group_id == group_id)
    offset = (page - 1) * page_size
    # Order by most recent first
    query = query.order_by(Student.created_at.desc())
    result = await db.execute(query.offset(offset).limit(page_size))
    students = result.scalars().all()

    count_query = select(func.count(Student.id)).where(year_filter)
    if status:
        count_query = count_query.where(Student.status == status)
    elif include_archived:
        count_query = count_query.where(Student.status != StudentStatus.DELETED)
    else:
        count_query = count_query.where(Student.status == StudentStatus.ACTIVE)
    if search:
        count_query = count_query.where(
            or_(
                Student.first_name.ilike(f"%{search}%"),
                Student.last_name.ilike(f"%{search}%"),
                Student.ampula.ilike(f"%{search}%"),
            )
        )
    if group_id:
        count_query = count_query.where(Student.group_id == group_id)
    count_result = await db.execute(count_query)
    total = count_result.scalar()

    return DataResponse(
        data=[serialize_student(s) for s in students],
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=(total + page_size - 1) // page_size,
        ),
    )

@router.get("/comprehensive-export", dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def export_comprehensive_student_data(
    db: Annotated[AsyncSession, Depends(get_db)],
    group_id: Optional[int] = Query(None, description="Filter by specific group"),
    status: Optional[str] = Query(None, description="Filter by student status (active, archived, etc.)"),
):
    """
    Export comprehensive student data to Excel file including:
    - Student information (name, ampula, phone, address, date of birth, status)
    - Group information

    Filters:
    - group_id: Filter students by specific group
    - status: Filter by student status (e.g., 'active', 'archived')
    """
    # Build student query with filters
    students_query = select(Student).options(selectinload(Student.group))

    if group_id:
        students_query = students_query.where(Student.group_id == group_id)

    if status:
        try:
            student_status = StudentStatus(status.lower())
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid student status: {status}")
        students_query = students_query.where(Student.status == student_status)

    students_result = await db.execute(
        students_query.order_by(Student.last_name.asc(), Student.first_name.asc())
    )
    students = students_result.scalars().all()
    active_students_count = sum(1 for s in students if s.status == StudentStatus.ACTIVE)
    inactive_students_count = sum(1 for s in students if s.status == StudentStatus.INACTIVE)

    # Prepare data for Excel
    student_data_list = []

    for student in students:
        group_name = student.group.name if student.group else "N/A"
        student_data_list.append({
            "student_id": student.id,
            "first_name": student.first_name,
            "last_name": student.last_name,
            "date_of_birth": student.date_of_birth.strftime("%Y-%m-%d"),
            "height": student.height,
            "weight": student.weight,
            "ampula": student.ampula or "N/A",
            "millati": student.millati or "N/A",
            "pnfl": student.pnfl,
            "phone": student.phone or "N/A",
            "address": student.address or "N/A",
            "status": student.status.value,
            "group": group_name,
        })

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Data"

    # Define header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Define headers
    headers = [
        "Student ID",
        "First Name",
        "Last Name",
        "Date of Birth",
        "Height",
        "Weight",
        "Ampula",
        "Millati",
        "PNFL",
        "Phone",
        "Address",
        "Status",
        "Group",
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    def _row_sort_key(row: dict) -> tuple:
        return (
            str(row.get("last_name", "")),
            str(row.get("first_name", "")),
            int(row.get("student_id", 0)),
        )

    student_data_list.sort(key=_row_sort_key)

    def _write_student_row(row_num: int, student_data: dict) -> None:
        ws.cell(row=row_num, column=1, value=student_data["student_id"])
        ws.cell(row=row_num, column=2, value=student_data["first_name"])
        ws.cell(row=row_num, column=3, value=student_data["last_name"])
        ws.cell(row=row_num, column=4, value=student_data["date_of_birth"])
        ws.cell(row=row_num, column=5, value=student_data["height"])
        ws.cell(row=row_num, column=6, value=student_data["weight"])
        ws.cell(row=row_num, column=7, value=student_data["ampula"])
        ws.cell(row=row_num, column=8, value=student_data["millati"])
        ws.cell(row=row_num, column=9, value=student_data["pnfl"])
        ws.cell(row=row_num, column=10, value=student_data["phone"])
        ws.cell(row=row_num, column=11, value=student_data["address"])
        ws.cell(row=row_num, column=12, value=student_data["status"])
        ws.cell(row=row_num, column=13, value=student_data["group"])

    current_row = 2
    for row in student_data_list:
        _write_student_row(current_row, row)
        current_row += 1

    # Adjust column widths
    column_widths = {
        'A': 12, 'B': 15, 'C': 15, 'D': 15, 'E': 10, 'F': 10,
        'G': 18, 'H': 18, 'I': 18, 'J': 15, 'K': 30, 'L': 12, 'M': 20
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Add summary row
    if student_data_list:
        summary_row = current_row + 1
        ws.cell(row=summary_row, column=1, value="TOTALS").font = Font(bold=True)

        # Add metadata
        metadata_row = summary_row + 2
        ws.cell(row=metadata_row, column=1, value=f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = Font(italic=True)
        ws.cell(row=metadata_row + 2, column=1, value=f"Total Students: {len(students)}").font = Font(italic=True)
        ws.cell(row=metadata_row + 4, column=1, value=f"Active Students: {active_students_count}").font = Font(italic=True)
        ws.cell(row=metadata_row + 5, column=1, value=f"Inactive Students: {inactive_students_count}").font = Font(italic=True)

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Generate filename
    group_suffix = f"_group{group_id}" if group_id else ""
    status_suffix = f"_{status}" if status else ""
    date_str = datetime.now().strftime('%Y%m%d')
    filename = f"comprehensive_student_data_{date_str}{group_suffix}{status_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Return as streaming response
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("", response_model=DataResponse[StudentRead], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def create_student(
    db: Annotated[AsyncSession, Depends(get_db)],
    student_payload: Annotated[StudentCreate, Depends(build_student_create_form)],
    photo: UploadFile | None = File(None, description="Optional student profile image"),
    passport: UploadFile | None = File(None, description="Optional student passport file"),
    extra_file: UploadFile | None = File(None, description="Optional extra student file"),
):
    try:
        if photo is not None:
            student_payload.photo_url = await upload_image_to_s3(photo, folder="student-profil-images")
        if passport is not None:
            student_payload.passport_key = await upload_private_file_to_s3(passport, folder="student-passports")
        if extra_file is not None:
            student_payload.extra_file_key = await upload_private_file_to_s3(extra_file, folder="student-extra-files")
    except Exception as exc:
        raise_student_file_upload_error(exc)

    student_payload.status = StudentStatus.ACTIVE

    existing_pnfl = await db.execute(select(Student).where(Student.pnfl == student_payload.pnfl))
    if existing_pnfl.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="PNFL already exists. Please use a unique PNFL")

    if student_payload.group_id:
        from app.models.domain import Group
        group_result = await db.execute(select(Group).where(Group.id == student_payload.group_id))
        group = group_result.scalar_one_or_none()
        if not group:
            raise HTTPException(status_code=404, detail=f"Group with ID {student_payload.group_id} not found")

        active_students_count = await db.execute(
            select(func.count(Student.id)).where(
                Student.group_id == student_payload.group_id,
                Student.status == StudentStatus.ACTIVE,
            )
        )
        current_count = active_students_count.scalar() or 0

        if current_count >= group.capacity:
            raise HTTPException(
                status_code=409,
                detail=f"Group '{group.name}' is at full capacity ({group.capacity} active students). "
                       f"Cannot add more active students."
            )

    student_data = student_payload.model_dump()
    student_data["archive_year"] = datetime.now().year
    student = Student(**student_data)
    db.add(student)
    await db.commit()
    await db.refresh(student)
    return DataResponse(data=serialize_student(student))

@router.get("/{student_id}", response_model=DataResponse[StudentRead], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_student(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    return DataResponse(data=serialize_student(student))


@router.get("/fullinfo/{student_id}", response_model=DataResponse[StudentFullInfo], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_student_full_info(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Get complete student information including:
    - Student details
    - Group
    - Coach (teacher)
    - Attendance records
    """
    from app.models.domain import Group
    from app.models.auth import User

    # Fetch student
    student_result = await db.execute(select(Student).where(Student.id == student_id))
    student = student_result.scalar_one_or_none()

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    # Fetch group and coach if student has a group
    group = None
    coach = None
    if student.group_id:
        group_result = await db.execute(select(Group).where(Group.id == student.group_id))
        group = group_result.scalar_one_or_none()

        if group and group.coach_id:
            coach_result = await db.execute(
                select(User).where(User.id == group.coach_id)
            )
            coach = coach_result.scalar_one_or_none()

    # Fetch attendance records
    attendances_result = await db.execute(
        select(Attendance).where(Attendance.student_id == student_id).order_by(Attendance.created_at.desc())
    )
    attendances = attendances_result.scalars().all()

    # Build the full info response
    from app.schemas.group import GroupRead
    from app.schemas.auth import UserRead

    full_info = StudentFullInfo(
        student=serialize_student(student),
        group=GroupRead.model_validate(group) if group else None,
        coach=UserRead.model_validate(coach) if coach else None,
        attendances=[AttendanceRead.model_validate(a) for a in attendances],
    )

    return DataResponse(data=full_info)


@router.patch("/{student_id}", response_model=DataResponse[StudentRead], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def update_student(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    student_update: Annotated[StudentUpdate, Depends(build_student_update_form)],
    photo: UploadFile | None = File(None, description="Optional student profile image"),
    passport: UploadFile | None = File(None, description="Optional student passport file"),
    extra_file: UploadFile | None = File(None, description="Optional extra student file"),
):
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    update_data = student_update.model_dump(exclude_unset=True)

    try:
        if photo is not None:
            update_data["photo_url"] = await upload_image_to_s3(photo, folder="student-profil-images")
        if passport is not None:
            update_data["passport_key"] = await upload_private_file_to_s3(passport, folder="student-passports")
        if extra_file is not None:
            update_data["extra_file_key"] = await upload_private_file_to_s3(extra_file, folder="student-extra-files")
    except Exception as exc:
        raise_student_file_upload_error(exc)

    # Prevent changing student's group once assigned
    if "group_id" in update_data and student.group_id is not None:
        if update_data["group_id"] != student.group_id:
            raise HTTPException(
                status_code=400,
                detail="Cannot change student's group. Students cannot be transferred between groups."
            )

    if "pnfl" in update_data and update_data["pnfl"] is not None:
        existing_pnfl = await db.execute(
            select(Student).where(Student.pnfl == update_data["pnfl"], Student.id != student_id)
        )
        if existing_pnfl.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="PNFL already exists. Please use a unique PNFL")

    # Check capacity when assigning to a group or changing status to ACTIVE
    target_group_id = update_data.get("group_id", student.group_id)
    target_status = update_data.get("status", student.status)

    # Determine if this update will make student ACTIVE in a group
    will_be_active_in_group = (
        target_group_id is not None and
        target_status == StudentStatus.ACTIVE and
        (student.group_id != target_group_id or student.status != StudentStatus.ACTIVE)
    )

    if will_be_active_in_group:
        from app.models.domain import Group
        group_result = await db.execute(select(Group).where(Group.id == target_group_id))
        group = group_result.scalar_one_or_none()

        if not group:
            raise HTTPException(status_code=404, detail=f"Group with ID {target_group_id} not found")

        # Count current ACTIVE students in target group (excluding this student).
        active_students_count = await db.execute(
            select(func.count(Student.id)).where(
                Student.group_id == target_group_id,
                Student.status == StudentStatus.ACTIVE,
                Student.id != student_id,
            )
        )
        current_count = active_students_count.scalar() or 0

        if current_count >= group.capacity:
            raise HTTPException(
                status_code=409,
                detail=f"Group '{group.name}' is at full capacity ({group.capacity} active students). "
                       f"Cannot add more active students."
            )
    elif "group_id" in update_data and update_data["group_id"] is not None:
        # Just validate group exists if only changing group
        from app.models.domain import Group
        group_result = await db.execute(select(Group).where(Group.id == update_data["group_id"]))
        if not group_result.scalar_one_or_none():
            raise HTTPException(status_code=404, detail=f"Group with ID {update_data['group_id']} not found")

    for field, value in update_data.items():
        setattr(student, field, value)

    await db.commit()
    await db.refresh(student)
    return DataResponse(data=serialize_student(student))


@router.post("/{student_id}/photo", response_model=DataResponse[StudentRead], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def upload_student_photo(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    photo: UploadFile = File(..., description="Student profile image"),
):
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    try:
        student.photo_url = await upload_image_to_s3(photo, folder="student-profil-images")
    except Exception as exc:
        raise_student_file_upload_error(exc)
    await db.commit()
    await db.refresh(student)
    return DataResponse(data=serialize_student(student))


@router.post("/{student_id}/passport", response_model=DataResponse[StudentRead], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def upload_student_passport(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    passport: UploadFile = File(..., description="Student passport file"),
):
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    try:
        student.passport_key = await upload_private_file_to_s3(passport, folder="student-passports")
    except Exception as exc:
        raise_student_file_upload_error(exc)
    await db.commit()
    await db.refresh(student)
    return DataResponse(data=serialize_student(student))


@router.post("/{student_id}/extra-file", response_model=DataResponse[StudentRead], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def upload_student_extra_file(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    extra_file: UploadFile = File(..., description="Student extra file"),
):
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    try:
        student.extra_file_key = await upload_private_file_to_s3(extra_file, folder="student-extra-files")
    except Exception as exc:
        raise_student_file_upload_error(exc)
    await db.commit()
    await db.refresh(student)
    return DataResponse(data=serialize_student(student))


@router.get("/files/{token}", include_in_schema=False)
async def get_student_file(token: str):
    payload = verify_private_file_token(token)
    key = payload["key"]

    try:
        file_bytes, content_type = await download_s3_object(key)
    except Exception as exc:
        raise HTTPException(status_code=404, detail=f"File not found: {str(exc)}") from exc

    filename = key.rsplit("/", 1)[-1]
    media_type = content_type or "application/octet-stream"
    disposition = "inline" if media_type.startswith("image/") or media_type == "application/pdf" else "attachment"
    headers = {"Content-Disposition": f'{disposition}; filename="{filename}"'}
    return StreamingResponse(iter([file_bytes]), media_type=media_type, headers=headers)


'''
@router.get("/{student_id}/transactions", response_model=DataResponse[list[TransactionRead]], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_student_transactions(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(Transaction).where(Transaction.student_id == student_id))
    transactions = result.scalars().all()
    return DataResponse(data=[TransactionRead.model_validate(t) for t in transactions])


'''

@router.get("/{student_id}/attendance", response_model=DataResponse[list[AttendanceRead]], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_student_attendance(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(Attendance).where(Attendance.student_id == student_id))
    attendance = result.scalars().all()
    return DataResponse(data=[AttendanceRead.model_validate(a) for a in attendance])


@router.get("/{student_id}/gatelogs", response_model=DataResponse[list[GateLogRead]], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_student_gatelogs(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(GateLog).where(GateLog.student_id == student_id))
    logs = result.scalars().all()
    return DataResponse(data=[GateLogRead.model_validate(l) for l in logs])


@router.get("/attendances/all", response_model=DataResponse[list[AttendanceRead]], dependencies=[Depends(require_permission(PERM_ATTENDANCE_VIEW))])
async def get_all_attendances(
    db: Annotated[AsyncSession, Depends(get_db)],
    from_date: Optional[date] = Query(None, description="Start date for filtering (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="End date for filtering (YYYY-MM-DD)"),
    group_id: Optional[int] = Query(None, description="Filter by group ID"),
    student_id: Optional[int] = Query(None, description="Filter by student ID"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=2000, description="Items per page"),
):
    """
    Get all student attendances with filters.

    This endpoint allows authorized users to view all attendance records
    with various filters including date range, group, and student.

    Filters:
    - from_date: Filter by session date (start)
    - to_date: Filter by session date (end)
    - group_id: Filter by specific group
    - student_id: Filter by specific student
    - page: Page number for pagination
    - page_size: Number of records per page (max 2000)

    Note: This is for viewing marked attendances (coach-created).
    Turnstile/gate attendance is handled separately.
    """
    from sqlalchemy.orm import selectinload

    # Build query for all attendances
    attendance_query = select(Attendance).options(
        selectinload(Attendance.student),
        selectinload(Attendance.session).selectinload(Session.group),
        selectinload(Attendance.marked_by)
    )

    # Apply date filters via session
    if from_date or to_date:
        attendance_query = attendance_query.join(Session)
        if from_date:
            attendance_query = attendance_query.where(Session.session_date >= from_date)
        if to_date:
            attendance_query = attendance_query.where(Session.session_date <= to_date)

    # Apply group filter via session
    if group_id:
        if not (from_date or to_date):  # Only join if not already joined
            attendance_query = attendance_query.join(Session)
        attendance_query = attendance_query.where(Session.group_id == group_id)

    # Apply student filter
    if student_id:
        attendance_query = attendance_query.where(Attendance.student_id == student_id)

    # Order by most recent first
    attendance_query = attendance_query.order_by(Attendance.created_at.desc())

    # Get total count for pagination
    count_query = select(func.count()).select_from(attendance_query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    # Apply pagination
    offset = (page - 1) * page_size
    attendance_query = attendance_query.offset(offset).limit(page_size)

    # Execute query
    attendances_result = await db.execute(attendance_query)
    attendances = attendances_result.scalars().all()

    return DataResponse(
        data=[AttendanceRead.model_validate(a) for a in attendances],
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=(total + page_size - 1) // page_size,
        ),
    )


@router.delete("/{student_id}", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def delete_student(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Soft delete a student by setting their status to DELETED.
    The student is not actually removed from the database.

    For permanent deletion, use DELETE /students/{student_id}/hard-delete
    """
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    # Soft delete: set status to DELETED instead of actually deleting
    student.status = StudentStatus.DELETED
    await db.commit()

    return DataResponse(data={"message": "Student deleted successfully"})


@router.delete("/{student_id}/hard-delete", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def hard_delete_student(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    **PERMANENT DELETION** - Remove student and all related data from database.

    **WARNING**: This action is irreversible!

    Deletes:
    - Student record
    - All attendance records (CASCADE)
    - All gate logs (CASCADE)

    Use this only when:
    - Student record was created by mistake
    - Duplicate entry needs to be removed
    - GDPR/data removal request

    For normal operations, use soft delete (DELETE /students/{student_id})
    """
    from sqlalchemy import delete as sql_delete

    # First get student name for response
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    student_name = f"{student.first_name} {student.last_name}"

    # HARD DELETE using raw SQL to avoid SQLAlchemy relationship loading issues
    # CASCADE will automatically delete:
    # - attendance, gate_logs
    await db.execute(sql_delete(Student).where(Student.id == student_id))
    await db.commit()

    return DataResponse(data={
        "message": "Student permanently deleted from database",
        "student_id": student_id,
        "student_name": student_name,
        "warning": "This action is irreversible. All related data has been removed."
    })


@router.post("/bulk-delete", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def bulk_delete_students(
    student_ids: list[int],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Soft delete multiple students by setting their status to DELETED.
    Students are not actually removed from the database.
    """
    if not student_ids:
        raise HTTPException(status_code=400, detail="No student IDs provided")

    deleted_count = 0
    errors = []

    for student_id in student_ids:
        try:
            result = await db.execute(select(Student).where(Student.id == student_id))
            student = result.scalar_one_or_none()

            if not student:
                errors.append({"student_id": student_id, "error": "Student not found"})
                continue

            # Soft delete: set status to DELETED instead of actually deleting
            student.status = StudentStatus.DELETED
            deleted_count += 1
        except Exception as e:
            errors.append({"student_id": student_id, "error": str(e)})

    await db.commit()

    return DataResponse(data={
        "message": f"Deleted {deleted_count} student(s)",
        "deleted_count": deleted_count,
        "total_requested": len(student_ids),
        "errors": errors if errors else None
    })

