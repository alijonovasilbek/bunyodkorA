import argparse
import asyncio
import json
import os
from datetime import date, datetime
from pathlib import Path
from typing import Any

import asyncpg
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TITLE = "Суғурталанган спортчилар рўйхати"
HEADERS = [
    "№",
    "Спортчилар исми шарифи",
    "Гувоҳнома ёки паспорт серияси",
    "Туғилган санаси",
    "Бир спортчи учун белгиланган суғурта пули (сўм)",
    "Бир спортчи учун белгиланган суғурта мукофоти (сўм)",
]
DEFAULT_OUTPUT = "sportchilar_export.xlsx"


def load_custom_fields(raw_value: Any) -> dict[str, Any]:
    if raw_value is None:
        return {}

    if isinstance(raw_value, dict):
        return raw_value

    if isinstance(raw_value, str):
        try:
            parsed = json.loads(raw_value)
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}

    return {}


def nested_get(data: dict[str, Any], *keys: str) -> Any:
    current: Any = data
    for key in keys:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def first_non_empty(*values: Any) -> str:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def build_full_name(custom_fields: dict[str, Any]) -> str:
    direct_name = first_non_empty(
        nested_get(custom_fields, "tarbiyalanuvchi", "fio"),
        nested_get(custom_fields, "student", "student_fio"),
        nested_get(custom_fields, "student_birth_certificate", "full_name"),
    )
    if direct_name:
        return direct_name

    parts = [
        nested_get(custom_fields, "student", "last_name"),
        nested_get(custom_fields, "student", "first_name"),
        nested_get(custom_fields, "student", "patronymic"),
    ]
    return " ".join(str(part).strip() for part in parts if part and str(part).strip())


def build_document_series(custom_fields: dict[str, Any]) -> str:
    return first_non_empty(
        nested_get(custom_fields, "tarbiyalanuvchi", "tugilganlik_guvohnoma"),
        nested_get(custom_fields, "student_birth_certificate", "series"),
        nested_get(custom_fields, "tarbiyalanuvchi", "pasport_seriya"),
        nested_get(custom_fields, "student", "passport_seriya"),
        nested_get(custom_fields, "student", "passport_series"),
        nested_get(custom_fields, "student", "passport_series_number"),
        nested_get(custom_fields, "buyurtmachi", "pasport_seriya"),
    )


def format_date_like_excel(value: date | datetime) -> str:
    date_value = value.date() if isinstance(value, datetime) else value
    return f"{date_value.day}/{date_value.month}/{date_value.year}"


def parse_date_string(value: str) -> str:
    raw = value.strip()
    if not raw:
        return ""

    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return format_date_like_excel(datetime.strptime(raw, fmt))
        except ValueError:
            continue
    return raw


def build_birth_value(custom_fields: dict[str, Any], student_date_of_birth: Any, contract_birth_year: Any) -> str:
    if isinstance(student_date_of_birth, (date, datetime)):
        return format_date_like_excel(student_date_of_birth)

    if isinstance(student_date_of_birth, str) and student_date_of_birth.strip():
        parsed = parse_date_string(student_date_of_birth)
        if parsed:
            return parsed

    custom_date = first_non_empty(
        nested_get(custom_fields, "student", "date_of_birth"),
        nested_get(custom_fields, "tarbiyalanuvchi", "tugilgan_sana"),
        nested_get(custom_fields, "tarbiyalanuvchi", "date_of_birth"),
        nested_get(custom_fields, "student_birth_certificate", "birth_date"),
    )
    if custom_date:
        return parse_date_string(custom_date)

    return first_non_empty(
        nested_get(custom_fields, "student", "birth_year"),
        nested_get(custom_fields, "tarbiyalanuvchi", "tugilganlik_yil"),
        contract_birth_year,
    )


def resolve_database_url(args: argparse.Namespace) -> str:
    explicit_db_url = (args.db_url or "").strip()
    if explicit_db_url:
        return explicit_db_url.replace("postgresql+asyncpg://", "postgresql://")

    env_database_url = os.getenv("DATABASE_URL", "").strip()
    has_manual_override = any(
        [
            args.db_host,
            args.db_port,
            args.db_name,
            args.db_user,
            args.db_password,
        ]
    )
    if env_database_url and not has_manual_override:
        return env_database_url.replace("postgresql+asyncpg://", "postgresql://")

    user = (args.db_user or os.getenv("POSTGRES_USER", "")).strip()
    password = (args.db_password or os.getenv("POSTGRES_PASSWORD", "")).strip()
    host = (args.db_host or os.getenv("POSTGRES_HOST", "localhost")).strip()
    port = str(args.db_port or os.getenv("POSTGRES_PORT", "5432")).strip()
    db_name = (args.db_name or os.getenv("POSTGRES_DB", "")).strip()

    if not all([user, password, db_name]):
        raise RuntimeError(
            "Database config topilmadi. `DATABASE_URL` yoki `POSTGRES_USER`, "
            "`POSTGRES_PASSWORD`, `POSTGRES_DB` env larini to'g'rilang."
        )

    return f"postgresql://{user}:{password}@{host}:{port}/{db_name}"


async def fetch_contract_rows(
    database_url: str,
    status: str | None,
    archive_year: int | None,
) -> list[asyncpg.Record]:
    connection = await asyncpg.connect(database_url)
    try:
        query = """
            SELECT
                c.id,
                c.contract_number,
                c.birth_year,
                c.custom_fields,
                s.date_of_birth
            FROM contracts c
            LEFT JOIN students s ON s.id = c.student_id
            WHERE ($1::text IS NULL OR c.status = $1)
              AND ($2::int IS NULL OR c.archive_year = $2)
            ORDER BY c.id
        """
        return await connection.fetch(query, status, archive_year)
    finally:
        await connection.close()


def style_worksheet(ws, data_end_row: int) -> None:
    ws.merge_cells("A1:F1")
    ws["A1"] = TITLE
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    thin_side = Side(style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    header_row = 3
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(4, data_end_row + 1):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col in (1, 4, 5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = {
        1: 6,
        2: 42,
        3: 22,
        4: 18,
        5: 22,
        6: 22,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = f"A3:F{max(data_end_row, 3)}"
    ws.freeze_panes = "A4"


def build_workbook(rows: list[dict[str, Any]], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sportchilar"

    current_row = 4
    for item in rows:
        worksheet.cell(row=current_row, column=1, value=item["index"])
        worksheet.cell(row=current_row, column=2, value=item["full_name"])
        worksheet.cell(row=current_row, column=3, value=item["document_series"])
        worksheet.cell(row=current_row, column=4, value=item["birth_value"])
        worksheet.cell(row=current_row, column=5, value="")
        worksheet.cell(row=current_row, column=6, value="")
        current_row += 1

    data_end_row = max(4, current_row - 1)
    style_worksheet(worksheet, data_end_row)
    workbook.save(output_path)


def transform_rows(records: list[asyncpg.Record]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []

    for index, record in enumerate(records, start=1):
        custom_fields = load_custom_fields(record["custom_fields"])
        result.append(
            {
                "index": index,
                "full_name": build_full_name(custom_fields),
                "document_series": build_document_series(custom_fields),
                "birth_value": build_birth_value(
                    custom_fields=custom_fields,
                    student_date_of_birth=record["date_of_birth"],
                    contract_birth_year=record["birth_year"],
                ),
            }
        )

    return result


async def main() -> None:
    parser = argparse.ArgumentParser(
        description="contracts.custom_fields dan sportchilar uchun Excel fayl yaratadi."
    )
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Excel fayl nomi")
    parser.add_argument(
        "--status",
        default="ACTIVE",
        help="Contract status filtri. Hammasini olish uchun `ALL` deb bering.",
    )
    parser.add_argument("--archive-year", type=int, default=None, help="Faqat bitta archive_year ni olish")
    parser.add_argument("--db-url", default=None, help="To'liq PostgreSQL URL")
    parser.add_argument("--db-host", default=None, help="DB host, masalan localhost")
    parser.add_argument("--db-port", type=int, default=None, help="DB port, masalan 5432")
    parser.add_argument("--db-name", default=None, help="DB nomi")
    parser.add_argument("--db-user", default=None, help="DB user")
    parser.add_argument("--db-password", default=None, help="DB password")
    args = parser.parse_args()

    load_dotenv()

    status_filter = None if args.status.upper() == "ALL" else args.status.upper()
    output_path = Path(args.output).resolve()
    database_url = resolve_database_url(args)

    records = await fetch_contract_rows(database_url, status_filter, args.archive_year)
    rows = transform_rows(records)
    build_workbook(rows, output_path)

    print(f"Excel tayyor: {output_path}")
    print(f"Jami sportchi: {len(rows)}")


if __name__ == "__main__":
    asyncio.run(main())
